# report.py

# Standard library
import base64
from datetime import datetime
import os
import re
import traceback
import uuid
import pandas as pd
import json
from pathlib import Path

# Third-party
import matplotlib
import pytz

from app.utils.funder_weighted_achievement import make_landscape_header_figure

matplotlib.use("Agg")  # Safe backend for servers
import matplotlib.pyplot as plt
from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from sqlalchemy import text

# App utilities
from app.routes.auth import login_required
from app.utils.database import get_db_engine, log_alert
from app.utils.funder_missing_plot import (
    add_full_width_footer_svg,
    create_funder_missing_figure,
)
from app.utils.one_bar_one_line import provider_portrait_with_target, use_ppmori
from app.utils.missing_classes_report import build_missing_classes_pdf
from app.utils.region_report import build_region_report_pdf
import app.utils.report_three_bar_landscape as r3  # kept for other report types
from app.utils.competency_icons import build_icon_reoprt 
REPORT_DIR = Path("/tmp/wsfl_reports")
REPORT_DIR.mkdir(parents=True, exist_ok=True)
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
# Blueprint
report_bp = Blueprint("report_bp", __name__)


# ----------------------------
# Logging helpers
# ----------------------------
def results_to_rows(results):
    """
    Normalise results into: List[Dict[str, Any]]
    Supports:
    - None / [] / ()
    - list[dict]
    - list[RowMapping] (SQLAlchemy mappings)
    - pandas.DataFrame
    """
    if results is None:
        return []

    # pandas DataFrame
    try:
        import pandas as pd
        if isinstance(results, pd.DataFrame):
            if results.empty:
                return []
            return results.to_dict(orient="records")
    except Exception:
        pass

    # SQLAlchemy RowMapping / mappings().all() already yields dict-like,
    # but just force to dict to be safe.
    if isinstance(results, (list, tuple)):
        if len(results) == 0:
            return []
        first = results[0]
        if isinstance(first, dict):
            return list(results)
        try:
            return [dict(r) for r in results]
        except Exception:
            return list(results)

    # single mapping object?
    try:
        return [dict(results)]
    except Exception:
        return []
    
def _user_ctx():
    """Common fields for structured-ish logs (kept simple)."""
    return {
        "user": session.get("user_email"),
        "role": session.get("user_role"),
        "admin": session.get("user_admin"),
        "path": request.path if request else None,
    }


def _safe_form_keys():
    """Avoid logging PII-heavy form contents; keys are enough for debugging."""
    try:
        return sorted(list(request.form.keys()))
    except Exception:
        return []


# ----------------------------
# Small helpers
# ----------------------------
def slugify_filename(label: str, fallback: str = "report") -> str:
    """Turn a human label into a filesystem-safe filename chunk."""
    label = (label or "").strip()
    if not label:
        return fallback
    label = label.replace("&", "and")
    label = re.sub(r"[^A-Za-z0-9\-_]+", "_", label)
    label = re.sub(r"_+", "_", label).strip("_")
    return label or fallback


def get_available_terms(nearest_year, nearest_term):
    options = [(nearest_year, nearest_term)]
    if nearest_term > 1:
        options.append((nearest_year, nearest_term - 1))
    else:
        options.append((nearest_year - 1, 4))
    return options

def _get_report_category():
    category = request.form.get("report_category", "visual")
    if session.get("user_role") != "ADM":
        category = "visual"
    return category

def log_report_run(
    engine,
    report_name,
    year,
    term,
    region_name,
    funder_id,
    provider_id,
    school_id,
    params_json,
    success=True,
    error_message=None,
):
    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    EXEC dbo.LogReportGeneration
                        @ReportName = :report_name,
                        @GeneratedByEmail = :email,
                        @CalendarYear = :year,
                        @Term = :term,
                        @RegionName = :region,
                        @FunderID = :funder_id,
                        @ProviderID = :provider_id,
                        @MOENumber = :school_id,
                        @ParametersJson = :params_json,
                        @Success = :success,
                        @ErrorMessage = :error
                """),
                {
                    "report_name": report_name,
                    "email": session.get("user_email"),
                    "year": year,
                    "term": term,
                    "region": region_name,
                    "funder_id": funder_id,
                    "provider_id": provider_id,
                    "school_id": school_id,
                    "params_json": params_json,
                    "success": 1 if success else 0,
                    "error": error_message,
                }
            )
    except Exception as e:
        current_app.logger.warning(f"⚠ Failed to log report: {e}")
        
# ---------- Download endpoints with log_alert ----------
@report_bp.route("/Reporting/download_pdf")
@login_required
def download_pdf():
    try:
        report_id = session.get("report_id")
        pdf_name = session.get("report_pdf_filename") or "report.pdf"

        # No report generated this session
        if not report_id:
            flash("No PDF report has been generated yet.", "warning")
            return redirect(url_for("report_bp.new_reports"))

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        # File missing on disk (restart / cleanup / expired)
        if not pdf_path.exists():
            flash("The PDF report has expired. Please run the report again.", "warning")
            return redirect(url_for("report_bp.new_reports"))

        return send_file(
            pdf_path,
            download_name=pdf_name,
            as_attachment=True,
            mimetype="application/pdf",
        )

    except Exception as e:
        err_text = f"/Reporting/download_pdf failed: {e}\n{traceback.format_exc()}"
        current_app.logger.exception("❌ Error in /Reporting/download_pdf")
        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=None,
                link=url_for("report_bp.download_pdf", _external=True),
                message=err_text[:4000],
            )
        except Exception:
            pass
        flash("An error occurred while downloading the PDF.", "danger")
        return redirect(url_for("report_bp.new_reports"))


@report_bp.route("/Reporting/download_png")
@login_required
def download_png():
    try:
        report_id = session.get("report_id")
        if not report_id:
            flash("No PNG report has been generated yet.", "warning")
            return redirect(url_for("report_bp.new_reports"))

        png_name = session.get("report_png_filename") or "report.png"
        png_path = REPORT_DIR / f"{report_id}.png"

        if not png_path.exists():
            flash("Report image has expired. Please re-run the report.", "warning")
            return redirect(url_for("report_bp.new_reports"))

        return send_file(
            png_path,
            download_name=png_name,
            as_attachment=True,
            mimetype="image/png",
        )
    except Exception as e:
        err_text = f"/Reporting/download_png failed: {e}\n{traceback.format_exc()}"
        current_app.logger.exception("❌ Error in /Reporting/download_png")
        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=None,
                link=url_for("report_bp.download_png", _external=True),
                message=err_text[:4000],
            )
        except Exception:
            pass
        flash("An error occurred while downloading the PNG.", "danger")
        return redirect(url_for("report_bp.new_reports"))


def _get_form_defaults():
    """
    Get year, term, and selected report type.

    - On GET: defaults come from session (nearest term/year)
    - On POST: overrides come from the submitted form
    """
    if request.method == "POST":
        year = int(request.form.get("year", session.get("nearest_year")))
        term = int(request.form.get("term", session.get("nearest_term")))
        report_type = request.form.get("report_option")
    else:
        year = session.get("nearest_year")
        term = session.get("nearest_term")
        report_type = None

    return year, term, report_type

def _persist_preview_for_existing_report(
    *,
    report_id: str,
    fig,
    selected_term: int,
    selected_year: int,
    selected_funder_name: str | None,
    base_label_prefix: str,
):
    png_path = REPORT_DIR / f"{report_id}.png"
    print(base_label_prefix)
    if(base_label_prefix==f"FunderWeightedAverageLYvsYTD"):
        fig.savefig(png_path, format="png", dpi=200, bbox_inches="tight", pad_inches=0)
        
    else: 
        fig.savefig(png_path, format="png", dpi=200)
        
    plt.close(fig)

    funder_chunk = slugify_filename(selected_funder_name or "Funder")
    base_label = f"{base_label_prefix}_{funder_chunk}_T{selected_term}_{selected_year}"

    session["report_id"] = report_id
    session["report_png_filename"] = f"{base_label}.png"
    session["report_pdf_filename"] = f"{base_label}.pdf"

    return base64.b64encode(png_path.read_bytes()).decode("ascii")

def _get_sticky_ids():
    """
    Return (provider_id, school_id) from the form if POST,
    otherwise (None, None).
    """
    if request.method == "POST":
        provider_id = request.form.get("provider_id") or None
        school_id = request.form.get("school_id") or None
        return provider_id, school_id
    return None, None


def _get_funder_name_from_role_or_form(role: str) -> str | None:
    """
    Decide which funder name to use based on role and request.

    - FUN: funder name comes from session.desc
    - ADM: funder name comes from the 'selected_funder' dropdown (on POST)
    - PRO: no funder name (None)
    """
    if role == "FUN":
        return session.get("desc")

    if role == "ADM" and request.method == "POST":
        return request.form.get("selected_funder") or None

    return None


def _get_request_type():
    """
    Return (is_ajax, action) based on the current request.

    - action comes from request.form["action"]
    - is_ajax is True if:
        * ajax=1 in the form, OR
        * X-Requested-With == 'fetch'
    """
    action = request.form.get("action")

    is_ajax = (
        request.form.get("ajax") == "1"
        or request.headers.get("X-Requested-With") == "fetch"
    )

    return is_ajax, action


def _validate_required_entities(
    selected_type,
    role,
    selected_provider_id,
    selected_school_id,
    funder_id,
    region_id,
    is_ajax,
    selected_term,
    selected_year,
):
    """
    Validate that required entities (provider/funder/school) exist
    for the selected report type.

    Returns a Flask response if validation fails.
    Returns None if validation passes.
    """
    needs_region  = selected_type in {
        "region_ly_vs_target","region_ytd",    "region_coverage_report",

    }
    if needs_region and not region_id:
        msg = "Please choose a region."

        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=session.get("user_id"),
                link=request.url,
                message="/Reports: region required but missing",
            )
        except Exception:
            pass
    # ---- Provider required ----
    needs_provider = selected_type in {
        "provider_ytd_vs_target",
        "provider_ytd_vs_target_vs_funder",
        "provider_missing_classes",
    }
    
    if needs_provider and not selected_provider_id:
        msg = "Please choose a provider."

        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=session.get("user_id"),
                link=request.url,
                message="/Reports: provider required but missing",
            )
        except Exception:
            pass

        if is_ajax:
            return jsonify({"ok": False, "error": msg}), 400

        flash(msg, "warning")
        return render_template(
            "reportingnew.html",
            role=role,
            funder_name=session.get("desc"),
            results=None,
            plot_payload=None,
            plot_png_b64=None,
            selected_term=selected_term,
            selected_year=selected_year,
            selected_type=selected_type,
            entities_url=url_for("api_bp.get_entities"),
            display=False,
            no_data_banner=None,
        )

    # ---- Funder required ----
    needs_funder = selected_type in {
        "funder_ytd_vs_target",
        "ly_funder_vs_ly_national_vs_target",
        "provider_ytd_vs_target_vs_funder",
        "funder_missing_data",
        "funder_student_count",
        "funder_progress_summary",
        "funder_teacher_review_summary",
        "funder_ytd_vs_funder_ly",
        "funder_missing_classes", 
    }

    if needs_funder and not funder_id:
        msg = "Please choose a funder."

        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=session.get("user_id"),
                link=request.url,
                message="/Reports: funder required but missing",
            )
        except Exception:
            pass

        if is_ajax:
            return jsonify({"ok": False, "error": msg}), 400

        flash(msg, "warning")
        return render_template(
            "reportingnew.html",
            role=role,
            funder_name=session.get("desc"),
            results=None,
            plot_payload=None,
            plot_png_b64=None,
            selected_term=selected_term,
            selected_year=selected_year,
            selected_type=selected_type,
            entities_url=url_for("api_bp.get_entities"),
            display=False,
            no_data_banner=None,
        )

    # ---- School required ----
    needs_school = selected_type in {"school_ytd_vs_target"}

    if needs_school and not selected_school_id:
        msg = "Please choose a school."

        try:
            log_alert(
                email=session.get("user_email"),
                role=session.get("user_role"),
                entity_id=session.get("user_id"),
                link=request.url,
                message="/Reports: school required but missing",
            )
        except Exception:
            pass

        if is_ajax:
            return jsonify({"ok": False, "error": msg}), 400

        flash(msg, "warning")
        return render_template(
            "reportingnew.html",
            role=role,
            funder_name=session.get("desc"),
            results=None,
            plot_payload=None,
            plot_png_b64=None,
            selected_term=selected_term,
            selected_year=selected_year,
            selected_type=selected_type,
            entities_url=url_for("api_bp.get_entities"),
            display=False,
            no_data_banner=None,
        )

    return None


def _execute_report(
    conn,
    selected_type,
    selected_year,
    selected_term,
    role,
    funder_id,
    region_id,
    selected_provider_id,
    selected_school_id,
    selected_funder_name,
    is_ajax,
):
    """
    Run the appropriate SQL/stored procedure and return:
        results, fig, no_data_banner, early_response
    """
    results = None
    fig = None
    no_data_banner = None
    early_response = None

    # 1) Funder YTD vs Target (data only)
    if selected_type == "funder_ytd_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetFunderNationalRates_All
                 @Term = :Term,
                 @CalendarYear = :CalendarYear;
            """
        )
        params = {"Term": selected_term, "CalendarYear": selected_year}
        res = conn.execute(sql, params)
        rows = res.mappings().all()

        if funder_id:
            funder_rows = [
                r
                for r in rows
                if (int(r.get("FunderID", 0) or 0) == funder_id)
                or (r.get("ResultType") == "WSNZ Target")
            ]

            if not funder_rows:
                unique_comps = {
                    (
                        r["CompetencyID"],
                        r["CompetencyDesc"],
                        r["YearGroupID"],
                        r["YearGroupDesc"],
                    )
                    for r in rows
                }
                funder_rows = []
                for cid, cdesc, yid, ydesc in unique_comps:
                    funder_rows.append(
                        {
                            "FunderID": funder_id,
                            "CompetencyID": cid,
                            "CompetencyDesc": cdesc,
                            "YearGroupID": yid,
                            "YearGroupDesc": ydesc,
                            "ResultType": "Funder Rate (YTD)",
                            "Rate": 0,
                            "StudentCount": 0,
                        }
                    )
                    funder_rows.append(
                        {
                            "FunderID": funder_id,
                            "CompetencyID": cid,
                            "CompetencyDesc": cdesc,
                            "YearGroupID": yid,
                            "YearGroupDesc": ydesc,
                            "ResultType": "Funder Student Count (YTD)",
                            "Rate": 0,
                            "StudentCount": 0,
                        }
                    )

            results = funder_rows
        else:
            results = rows

        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "national_competency_icons":
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        # build the PDF
        preview_fig  = build_icon_reoprt(
            out_pdf_path=pdf_path,
            footer_svg=Path(current_app.static_folder) / "footer.svg",
            dpi=300,
            footer_height_frac=0.10,
            term = selected_term,
            year = selected_year,
        )

        results = None
        fig = preview_fig
    # 2) National LY vs National YTD vs Target
    elif selected_type == "national_ly_vs_national_ytd_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC GetNationalRates
                @CalendarYear = :CalendarYear,
                @Term         = :Term;
            """
        )
        params = {"CalendarYear": selected_year, "Term": selected_term}
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "national_ytd_kaiako":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC GetNationalRates_Kaiako
                @CalendarYear = :CalendarYear,
                @Term         = :Term;
            """
        )
        params = {"CalendarYear": selected_year, "Term": selected_term}
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "funder_weighted_average":
        engine = get_db_engine()
        with engine.connect() as connection:
            result = connection.execute(
                text(
                    "EXEC GetFunderYearGroupSummary_StudentWeighted_TY_LY_WithTrend "
                    ":CalendarYear, :Term"
                ),
                {
                    "CalendarYear": selected_year,
                    "Term": selected_term,
                },
            )
            df = pd.DataFrame(result.fetchall(), columns=result.keys()) 
            df = df.loc[
            df["YearGroupDesc"] == "All Year Groups",
                ["Funder", "TY_AllYGsRate", "LY_AllYGsRate"],
            ].copy()
            fig =make_landscape_header_figure(df=df,
                                         title = "Funder Weighted Achievement Summary",
                                        subtitle = f"YTD (Term {selected_term}, {selected_year}) vs LY (Full Year)",
                                        term = selected_term, calendaryear=selected_year
                                        
                                        )
            
    elif selected_type == "region_coverage_report":

        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        preview_fig, meta = build_region_report_pdf(
            conn=conn,
            region_name=region_id,
            calendar_year=selected_year,
            term=selected_term,
            out_pdf_path=pdf_path,
            draw_key=False,
            footer_svg=Path(current_app.static_folder) / "footer.svg",
            dpi=300,
            page_size="A4",
            orientation="portrait",
            fonts_dir="app/static/fonts",
        )

        current_app.logger.info(
            "📄 region_report PDF pages=%s region=%s",
            meta.get("pages"),
            region_id,
        )

        results = None
        fig = preview_fig
            
    elif selected_type == "region_ly_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC GetRegionalCouncilRates
                @CalendarYear = :CalendarYear,
                @Term         = :Term,
                @Region = :Region;
            """
        )
        params = {"CalendarYear": 2025, "Term": 2,"Region":region_id}
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "region_ytd":
        
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC [GetRegionalCouncilRates_kaiako]
                @CalendarYear = :CalendarYear,
                @Term         = :Term,
                @Region = :Region;
            """
        )
        params = {"CalendarYear": selected_year, "Term": selected_term,"Region":region_id}
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    # ✅ IMPORTANT: return list-of-row-mappings (NOT DataFrame)
    elif selected_type == "funder_targets_counts":
        sql = text("SET NOCOUNT ON; EXEC GetFunderTargetsCounts;")
        results = conn.execute(sql).mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)

    elif selected_type == "national_ytd_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC GetNationalRates
                @CalendarYear = :CalendarYear,
                @Term         = :Term;
            """
        )
        params = {"CalendarYear": selected_year, "Term": selected_term}
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)

    # 3) Funder Missing Data (builds fig here)
    elif selected_type == "funder_missing_data":
        threshold = 0.25
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC FlaskGetSchoolSummaryByFunder
                @CalendarYear = :CalendarYear,
                @Term         = :Term,
                @Threshold    = :Threshold,
                @Email        = :Email,
                @FunderName = :f;
            """
        )
        params = {
            "CalendarYear": selected_year,
            "Term": selected_term,
            "Threshold": threshold,
            "Email": session.get("user_email"),
            "f": selected_funder_name,
        }
        res = conn.execute(sql, params)
        rows = res.mappings().all()
        results = rows

        df_all = pd.DataFrame(rows)
        
        if df_all.empty:
            msg = f"No data found for funder: {selected_funder_name}"
            if is_ajax:
                early_response = (jsonify({"ok": False, "error": msg}), 400)
            else:
                flash(msg, "warning")
                early_response = redirect(url_for("report_bp.new_reports"))
            return results, fig, no_data_banner, early_response

        fig = create_funder_missing_figure(
            df_all=df_all,
            funder_name=selected_funder_name,
            term=selected_term,
            calendaryear=selected_year,
            threshold=threshold,
            debug=False,
        )

        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)

    # 4) Funder LY vs National LY vs Target (data only)
    elif selected_type == "ly_funder_vs_ly_national_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetFunderNationalRates_All
                @Term = :Term,
                @CalendarYear = :CalendarYear;
            """
        )
        params = {"Term": 2, "CalendarYear": 2025}
        res = conn.execute(sql, params)

        raw_rows = res.mappings().all()
        filtered_rows = []
        for r in raw_rows:
            funder_matches = not funder_id or int(r.get("FunderID", 0) or 0) == funder_id
            keep = (
                (funder_matches and r.get("ResultType") == "Funder Rate (YTD)")
                or r.get("ResultType") == "WSNZ Target"
                or r.get("ResultType") == "National Rate (YTD)"
            )
            if not keep:
                continue

            d = dict(r)
            if d.get("ResultType") == "National Rate (YTD)":
                d["ResultType"] = "National Rate (LY)"
            elif d.get("ResultType") == "Funder Rate (YTD)":
                d["ResultType"] = "Funder Rate (LY)"

            filtered_rows.append(d)

        results = filtered_rows
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "funder_ytd_vs_funder_ly":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetFunderNationalRates_All
                @Term = :Term,
                @CalendarYear = :CalendarYear;
            """
        )
        params = {"Term": selected_term, "CalendarYear": selected_year}
        res = conn.execute(sql, params)

        raw_rows = res.mappings().all()
        filtered_rows = []
        for r in raw_rows:
            funder_matches = not funder_id or int(r.get("FunderID", 0) or 0) == funder_id
            keep = (
                (funder_matches and r.get("ResultType") == "Funder Rate (YTD)")
                or (funder_matches and r.get("ResultType") == "Funder Rate (LY)")
            )
            if not keep:
                continue

            d = dict(r)
            
            filtered_rows.append(d)

        results = filtered_rows
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    # 5) Provider vs Funder (data only)
    elif selected_type == "provider_ytd_vs_target_vs_funder":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetProviderNationalRates
                @Term         = :Term,
                @CalendarYear = :CalendarYear,
                @ProviderID   = :ProviderID,
                @FunderID     = :FunderID;
            """
        )
        params = {
            "Term": selected_term,
            "CalendarYear": selected_year,
            "ProviderID": int(selected_provider_id),
            "FunderID": int(funder_id) if funder_id is not None else None,
        }
        res = conn.execute(sql, params)
        rows = res.mappings().all()

        if len(rows) == 0:
            res2 = conn.exec_driver_sql(
                "SET NOCOUNT ON; EXEC dbo.GetProviderNationalRates @Term=?, @CalendarYear=?, @ProviderID=?, @FunderID=?",
                (
                    selected_term,
                    selected_year,
                    int(selected_provider_id),
                    funder_id if funder_id is not None else None,
                ),
            )
            if getattr(res2, "cursor", None) and res2.cursor.description:
                cols = [d[0] for d in res2.cursor.description]
                rows = [dict(zip(cols, row)) for row in res2.fetchall()]

        results = rows
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "funder_missing_classes":
    # --- PDF report (multi-page) for FUNDERS ---
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        base_label_prefix = "funder-missing_classes"
        threshold = 0.25

        # Build PDF + return a preview fig for UI
        preview_fig, meta = build_missing_classes_pdf(
            conn=conn,
            calendar_year=selected_year,
            term=selected_term,
            funder_id=int(funder_id) if funder_id is not None else None,
            provider_id=None,  # ✅ IMPORTANT: funder-wide
            threshold=threshold,
            email=session.get("user_email"),
            out_pdf_path=pdf_path,
            footer_svg=Path(current_app.static_folder) / "footer.svg",
            dpi=300,
            page_size="A4",
            orientation="portrait",
            fonts_dir="app/static/fonts",
            max_rows_per_page=25,
        )

        current_app.logger.info(
            "📄 funder_missing_classes PDF pages=%s rows=%s raw=%s mode=%s funder_id=%s",
            meta.get("pages"),
            meta.get("rows_display"),
            meta.get("rows_raw"),
            meta.get("mode"),
            funder_id,
        )

        results = None
        fig = preview_fig
    elif selected_type == "provider_missing_classes":
        # --- PDF report (multi-page) ---
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        # If you want the filename to say provider-missing_classes
        provider_name = (request.form.get("provider_name") or "").strip() or "Provider"
        base_label_prefix = "provider-missing_classes"

        # threshold: match your stored proc default usage
        threshold = 0.25

        # Build PDF + return a preview fig for UI
        preview_fig, meta = build_missing_classes_pdf(
            conn=conn,
            calendar_year=selected_year,
            term=selected_term,
            funder_id=int(funder_id) if funder_id is not None else None,
            provider_id=int(selected_provider_id),
            threshold=threshold,
            email=session.get("user_email"),
            out_pdf_path=pdf_path,
            footer_svg=Path(current_app.static_folder) / "footer.svg",
            dpi=300,
            page_size="A4",
            orientation="portrait",
            fonts_dir="app/static/fonts",
            max_rows_per_page=25,
        )

        # optional: log meta for debugging
        current_app.logger.info(
            "📄 provider_missing_classes PDF pages=%s rows=%s raw=%s mode=%s",
            meta.get("pages"),
            meta.get("rows_display"),
            meta.get("rows_raw"),
            meta.get("mode"),
        )

        # results are not used for this report type
        results = None
        fig = preview_fig
    # 6) Provider YTD vs Target (data only)
    elif selected_type == "provider_ytd_vs_target":
        if role == "ADM":
            funder_id = None

        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetProviderNationalRates
                 @Term         = :Term,
                 @CalendarYear = :CalendarYear,
                 @ProviderID   = :ProviderID,
                 @FunderID     = :FunderID;
            """
        )
        params = {
            "Term": selected_term,
            "CalendarYear": selected_year,
            "ProviderID": int(selected_provider_id),
            "FunderID": int(funder_id) if funder_id is not None else None,
        }
        res = conn.execute(sql, params)
        rows = res.mappings().all()
        results = rows

        current_app.logger.info("🧪 Rows fetched: %d", len(rows))

    # 7) School YTD vs Target (data only)
    elif selected_type == "school_ytd_vs_target":
        sql = text(
            """
            SET NOCOUNT ON;
            EXEC dbo.GetSchoolNationalRates
                 @CalendarYear = :CalendarYear,
                 @Term         = :Term,
                 @MoeNumber    = :MoeNumber;
            """
        )
        params = {
            "CalendarYear": selected_year,
            "Term": selected_term,
            "MoeNumber": int(selected_school_id),
        }
        res = conn.execute(sql, params)
        results = res.mappings().all()
        current_app.logger.info("🔎 rows=%d | type=%s", len(results or []), selected_type)
    elif selected_type == "funder_student_count":
        from app.utils.funder_student_counts import build_funder_student_counts_pdf
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)
        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id  # <-- important

        pdf_path = REPORT_DIR / f"{report_id}.pdf"
        footer_png = Path(current_app.static_folder) / "footer.png"

        preview_fig, meta = build_funder_student_counts_pdf(
            conn=conn,
            funder_id=funder_id,
            term = selected_term,
            year = selected_year,
            out_pdf_path=pdf_path,
            footer_png=footer_png,
        )

        results = None
        fig = preview_fig
    elif selected_type == "funder_progress_summary":
        from app.utils.funder_summary import build_funder_progress_summary_pdf

        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        preview_fig, meta = build_funder_progress_summary_pdf(
            conn=conn,
            funder_id=funder_id,
            from_year=selected_year,
            threshold=0.2,
            out_pdf_path=pdf_path,
            footer_png=None,
        )

        results = None
        fig = preview_fig

    elif selected_type == "funder_teacher_review_summary":
        from app.utils.teacher_assessment import build_funder_teacher_assessment_summary_pdf

        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        report_id = session.get("report_id") or uuid.uuid4().hex
        session["report_id"] = report_id

        pdf_path = REPORT_DIR / f"{report_id}.pdf"

        preview_fig, meta = build_funder_teacher_assessment_summary_pdf(
            conn=conn,
            funder_id=funder_id,
            out_pdf_path=pdf_path,
            footer_png=None,
            rows_per_page=30,
            dpi=300,
            page_size="A3",
            orientation="portrait",
            fonts_dir="app/static/fonts",
        )

        results = None
        fig = preview_fig

    else:
        results = None

    return results, fig, no_data_banner, early_response

def _build_figure_from_results(
    selected_type,
    results,
    selected_term,
    selected_year,
    funder_id,
    region_id,
    selected_funder_name,
    selected_provider_id,
    selected_school_id,
    rows,
):
    """Build a Matplotlib figure from the result rows. Returns (fig, no_data_banner)."""
    fig = None
    no_data_banner = None

    if not rows:
        return None, None

    if selected_type == "provider_ytd_vs_target_vs_funder":
        vars_to_plot = ["Provider Rate (YTD)", "Funder Rate (YTD)", "WSNZ Target"]
        colors_dict = {
            "Provider Rate (YTD)": "#2EBDC2",
            "WSNZ Target": "#356FB6",
            "Funder Rate (YTD)": "#BBE6E9",
        }

        provider_display_name = (
            request.form.get("provider_name")
            or next(
                (
                    r.get("ProviderName") or r.get("Provider")
                    for r in (rows or [])
                    if r.get("ProviderName") or r.get("Provider")
                ),
                None,
            )
            or f"Provider {selected_provider_id}"
        )
        funder_display_name = selected_funder_name or "Funder"
        title_text = f"{provider_display_name} & {funder_display_name}"

        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            funder_id=funder_id,
            rows=rows,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            funder_name=title_text,
        )
    elif selected_type == "funder_ytd_vs_funder_ly":
        vars_to_plot = ["Funder Rate (YTD)", "Funder Rate (LY)"]
        colors_dict = {
            "Funder Rate (YTD)": "#2EBDC2",
            "Funder Rate (LY)": "#BBE6E9",
        }
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            funder_id=funder_id or 0,
            rows=rows,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            funder_name=selected_funder_name,
        )
    elif selected_type == "ly_funder_vs_ly_national_vs_target":
        vars_to_plot = ["National Rate (LY)", "Funder Rate (LY)", "WSNZ Target"]
        colors_dict = {
            "Funder Rate (LY)": "#2EBDC2",
            "WSNZ Target": "#356FB6",
            "National Rate (LY)": "#BBE6E9",
        }
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            funder_id=funder_id or 0,
            rows=rows,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            funder_name=selected_funder_name,
        )

    elif selected_type == "national_ly_vs_national_ytd_vs_target":
        vars_to_plot = ["National Rate (LY)", "National Rate (YTD)", "WSNZ Target"]
        colors_dict = {
            "National Rate (YTD)": "#2EBDC2",
            "WSNZ Target": "#356FB6",
            "National Rate (LY)": "#BBE6E9",
        }
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            rows=rows,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            funder_id=None,
        )
    elif selected_type == "national_ytd_kaiako":
        vars_to_plot = ["National Instructor Rate (YTD)", "National Rate (YTD)", "National Kaiako Rate (YTD)"]
        colors_dict = {
            "National Instructor Rate (YTD)": "#2EBDC2",
            "National Rate (YTD)": "#356FB6",
            "National Kaiako Rate (YTD)": "#BBE6E9",
        }
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            rows=rows,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            funder_id=None,
        )
    elif selected_type == "national_ytd_vs_target":
        fig = provider_portrait_with_target(
            rows,
            term=selected_term,
            year=selected_year,
            mode="national",
            subject_name="",
            title="National YTD vs WSNZ Target",
        )

    elif selected_type == "provider_ytd_vs_target":
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        subject_name = (request.form.get("provider_name") or "").strip() or "Unknown Provider"

        def _is_provider_row(r):
            return str(r.get("ResultType", "")).lower().startswith("provider rate")

        provider_rows = [r for r in rows if _is_provider_row(r)]
        filtered_rows = rows
        if not provider_rows:
            no_data_banner = (
                f"No YTD provider data found for {subject_name} "
                f"(Term {selected_term}, {selected_year}). Showing national/target series only."
            )
            filtered_rows = [r for r in rows if not _is_provider_row(r)]

        chart_title = f"{subject_name} — YTD vs Target (Term {selected_term}, {selected_year})"

        fig = provider_portrait_with_target(
            filtered_rows,
            term=selected_term,
            year=selected_year,
            mode="provider",
            subject_name=subject_name,
            title=chart_title,
        )

        if no_data_banner and fig is not None:
            try:
                ax = fig.gca()
                ax.annotate(
                    no_data_banner,
                    xy=(0.5, 1.0),
                    xycoords="axes fraction",
                    ha="center",
                    va="bottom",
                    fontsize=10,
                    fontweight="bold",
                    color="white",
                    bbox=dict(boxstyle="round,pad=0.4", fc="#1a427d", ec="#1a427d"),
                )
            except Exception as _e:
                current_app.logger.info("⚠️ Could not annotate no-data banner: %s", _e)

    elif selected_type == "funder_ytd_vs_target":
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        fig = provider_portrait_with_target(
            rows,
            term=selected_term,
            year=selected_year,
            mode="funder",
            subject_name=selected_funder_name,
            title=f"{selected_funder_name or 'Funder'} YTD vs Target",
        )
    elif selected_type == "region_ly_vs_target":
        region_label = (request.form.get("region_name") or "").strip() or "Region"
        print(rows)
        fig = provider_portrait_with_target(
            rows,
            term=selected_term,
            year=selected_year,
            mode="region",
            subject_name=region_label,
            caption = (
                f"{region_label} Competency Rate for Term 3, 2024 - Term 2, 2025 | "
                f"Generated {datetime.now(pytz.timezone('Pacific/Auckland')).strftime('%d %b %Y, %I:%M %p')}"
            ),
            title=f"{region_label} Last Year Result vs WSNZ Target",
            bar_series="ytd",
        )
    elif selected_type == "region_ytd":
        vars_to_plot = ["Region Rate (YTD)", "Region Kaiako-Led Rate (YTD)", "Region Instructor-Led Rate (YTD)"]
        colors_dict = {
            "Region Instructor-Led Rate (YTD)": "#2EBDC2",
            "Region Rate (YTD)": "#356FB6",
            "Region Kaiako-Led Rate (YTD)": "#BBE6E9",
        }
        region_name = {
            r["RegionalCouncil"]
            for r in rows
            if r.get("RegionalCouncil") is not None
        }
        region_name = next(iter(region_name), None) 
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            rows=rows,
            funder_id = None,
            vars_to_plot=vars_to_plot,
            colors_dict=colors_dict,
            region_name=region_name
        )
    elif selected_type == "school_ytd_vs_target":
        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        school_name = request.form.get("school_name") or next(
            (r.get("SchoolName") for r in rows if r.get("SchoolName")), None
        )

        fig = provider_portrait_with_target(
            rows,
            term=selected_term,
            year=selected_year,
            mode="school",
            subject_name=school_name,
            title=f"{school_name or 'School'} YTD vs WSNZ Target",
        )

    elif selected_type == "funder_targets_counts":
        from app.utils.funder_targets_counts_report import build_funder_targets_counts_figure

        try:
            use_ppmori("app/static/fonts")
        except Exception as font_e:
            current_app.logger.info("⚠️ font setup skipped: %s", font_e)

        # IMPORTANT: pass normalised rows + let outer code add footer once
        fig, meta = build_funder_targets_counts_figure(
            rows,
            footer_svg=None,          # avoid double-footer
            fonts_dir="app/static/fonts",
            title="Funder Counts (Target vs Actual)",
        )

    elif selected_type in {"funder_student_count", "funder_progress_summary", "funder_teacher_review_summary"}:
        return None, None

    else:
        fig = r3.create_competency_report(
            term=selected_term,
            year=selected_year,
            funder_id=funder_id or 0,
            rows=rows,
            vars_to_plot=r3.vars_to_plot,
            colors_dict=r3.colors_dict,
            funder_name=selected_funder_name,
        )

    # Add footer once here for all normal figures (including funder_targets_counts),
    # BUT only if fig exists.
    if fig is not None and selected_type not in {"funder_weighted_average","provider_missing_classes", "funder_missing_classes", "region_coverage_report", "national_competency_icons",}:

        c = "#1a427d" if selected_type == "funder_missing_data" else "#1a427d40"
        try:
            footer_svg = os.path.join(current_app.static_folder, "footer.svg")
            add_full_width_footer_svg(
                fig,
                footer_svg,
                bottom_margin_frac=0.0,
                max_footer_height_frac=0.20,
                col_master=c,
            )
        except Exception as footer_e:
            current_app.logger.info("⚠ Could not add footer to figure: %s", footer_e)

    return fig, no_data_banner

def _persist_figure_and_session(
    fig,
    selected_type,
    selected_term,
    selected_year,
    selected_funder_name,
    selected_provider_id,
    selected_school_id,
    results,
):
    report_id = uuid.uuid4().hex

    png_path = REPORT_DIR / f"{report_id}.png"
    pdf_path = REPORT_DIR / f"{report_id}.pdf"
    if(selected_type != "funder_weighted_average"):
        fig.savefig(png_path, format="png", dpi=200)
        fig.savefig(pdf_path, format="pdf")
    else:
        fig.savefig(png_path, format="png", dpi=200, bbox_inches="tight", pad_inches=0)
        fig.savefig(pdf_path, format="pdf", bbox_inches="tight", pad_inches=0)
    plt.close(fig)

    provider_name = (request.form.get("provider_name") or "").strip()
    school_name = (request.form.get("school_name") or "").strip()

    # Normalise results safely for name picking (no DataFrame truthiness)
    rows = results_to_rows(results)

    if not provider_name and rows:
        provider_name = next(
            (
                (r.get("ProviderName") or r.get("Provider") or r.get("ProviderDesc"))
                for r in rows
                if r.get("ProviderName") or r.get("Provider") or r.get("ProviderDesc")
            ),
            "",
        )

    if not school_name and rows:
        school_name = next((r.get("SchoolName") for r in rows if r.get("SchoolName")), "")
    add_term = True
    if selected_type == "funder_missing_data":
        base_label = f"MissingData_{selected_funder_name or 'Funder'}"
        add_term = False
    elif selected_type=="region_coverage_report":
        region_label = (request.form.get("region_name") or "Region").strip()
        base_label = f"Region_full_summary_{region_label}"
        add_term = False
    elif selected_type == "funder_ytd_vs_target":
        base_label = f"FunderYTDvsTarget_{selected_funder_name or 'Funder'}"
    elif selected_type == "ly_funder_vs_ly_national_vs_target":
        base_label = f"FunderLY_vs_National_vs_Target_{selected_funder_name or 'Funder'}"
    elif selected_type == "provider_ytd_vs_target_vs_funder":
        base_label = f"ProviderVsFunder_{provider_name or 'Provider'}_{selected_funder_name or 'Funder'}"
    elif selected_type == "provider_ytd_vs_target":
        base_label = f"ProviderYTDvsTarget_{provider_name or 'Provider'}"
    elif selected_type =="provider_missing_classes":
        base_label = f"ProviderMissingClasses_{provider_name or 'Provider'}"
    elif selected_type == "funder_missing_classes":
        base_label = f"FunderMissingClasses_{selected_funder_name or 'Funder'}"
    elif selected_type == "school_ytd_vs_target":
        base_label = f"SchoolYTDvsTarget_{school_name or f'MOE_{selected_school_id}'}"
    elif selected_type == "national_ly_vs_national_ytd_vs_target":
        base_label = "NationalLYvsNationalYTDvsTarget"
    elif selected_type == "national_ytd_kaiako":
        base_label = "NationalYTD_KaiakoVsInstructor"
    elif selected_type == "national_ytd_vs_target":
        base_label = "NationalYTDvsTarget"
    elif selected_type == "funder_weighted_average":
        base_label = "FunderWeightedAverageLYvsYTD"
    elif selected_type == "funder_targets_counts":
        base_label = f"FunderTargetsCounts"
        add_term = False
    elif selected_type == "national_competency_icons":
        base_label = "NationalCompetencyIcons"
    elif selected_type == "region_ly_vs_target":
        region_label = (request.form.get("region_name") or "Region").strip()
        base_label = f"RegionLYvsTarget_{region_label}"
        add_term = False
    elif selected_type == "region_ytd":
        region_label = (request.form.get("region_name") or "Region").strip()
        base_label = f"RegionYTD_Kaiako_vs_Instructor_{region_label}"
        add_term = False
    elif selected_type == "funder_ytd_vs_funder_ly":
        base_label = f"FunderLYvsFunderTY_{selected_funder_name}"
    elif selected_type == "region_coverage_report":
        region_label = (request.form.get("region_name") or "Region").strip()
        base_label = f"RegionReport_{region_label}"
        add_term = False
    else:
        base_label = f"Report_{selected_type or 'Unknown'}"

    base_label = slugify_filename(base_label, fallback="WSFL_Report")
    print(add_term)
    if add_term:
        base_label = f"{base_label}_T{selected_term}_{selected_year}"

    session["report_id"] = report_id
    session["report_png_filename"] = f"{base_label}.png"
    session["report_pdf_filename"] = f"{base_label}.pdf"

    plot_png_b64 = base64.b64encode(png_path.read_bytes()).decode("ascii")
    return plot_png_b64
_table_counter = 1

def _convert_to_table(ws, table_name_prefix="Table"):
    global _table_counter

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 2:
        return

    # Excel table names:
    # - must start with a letter or underscore
    # - can only contain letters, numbers, underscores
    safe_prefix = re.sub(r"[^A-Za-z0-9_]", "_", str(table_name_prefix))
    if not safe_prefix:
        safe_prefix = "Table"
    if not re.match(r"^[A-Za-z_]", safe_prefix):
        safe_prefix = f"T_{safe_prefix}"

    end_col = get_column_letter(max_col)
    ref = f"A1:{end_col}{max_row}"

    table_name = f"{safe_prefix}_{_table_counter}"
    _table_counter += 1

    table = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(
        name="TableStyleLight8",
        showRowStripes=True,
        showColumnStripes=False
    )

    table.tableStyleInfo = style
    ws.add_table(table)
@report_bp.route("/Reports", methods=["GET", "POST"])
@login_required
def new_reports():
    ctx = _user_ctx()
    role = session.get("user_role")  # "ADM", "FUN", or "PRO"

    current_app.logger.info(
        "📄 /Reports hit | method=%s role=%s user=%s admin=%s",
        request.method,
        role,
        ctx["user"],
        ctx["admin"],
    )

    if session.get("user_admin") != 1:
        return (
            render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403,
            ),
            403,
        )

    engine = get_db_engine()

    # --------- UI state (defaults shown in form) ----------
    selected_year, selected_term, selected_type = _get_form_defaults()
    selected_provider_id, selected_school_id = _get_sticky_ids()

    selected_funder_name = _get_funder_name_from_role_or_form(role)

    if role == "PRO":
        selected_provider_id = selected_provider_id or session.get("id")
        current_app.logger.debug("PRO effective provider_id=%s", selected_provider_id)
    elif role == "FUN":
        current_app.logger.debug("FUN default funder_name(session.desc)=%s", selected_funder_name)
    
    current_app.logger.debug(
        "📥 initial ids | provider_id=%s school_id=%s | year=%s term=%s type=%s | form_keys=%s",
        selected_provider_id,
        selected_school_id,
        selected_year,
        selected_term,
        selected_type,
        _safe_form_keys(),
    )

    results = None
    plot_payload = None
    plot_png_b64 = None
    no_data_banner = None
    display = False

    is_ajax, action = _get_request_type()
    report_category = _get_report_category()
    excel_report_type = request.form.get("excel_report_option")

    if request.method == "POST" and action == "download_excel":
        if role != "ADM":
            flash("You are not authorised to download Excel exports.", "danger")
            return redirect(url_for("report_bp.new_reports"))

        try:
            excel_report_type = request.form.get("excel_report_option")
            selected_year = int(request.form.get("year", session.get("nearest_year")))
            selected_term = int(request.form.get("term", session.get("nearest_term")))
            selected_provider_id = request.form.get("provider_id") or None
            selected_school_id = request.form.get("school_id") or None
            selected_funder_name = request.form.get("funder_name") or None

            funder_id = None
            region_id = request.form.get("region_name") or None

            def _safe_sheet_name(name: str, fallback: str = "Sheet") -> str:
                bad = ['\\', '/', '*', '?', ':', '[', ']']
                name = str(name or fallback)
                for ch in bad:
                    name = name.replace(ch, "-")
                name = name.strip()
                return name[:31] or fallback

            def _autosize_and_freeze(ws):
                ws.freeze_panes = "A2"
                for col_cells in ws.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        val = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(val))
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
            
            with engine.connect() as conn:
                if selected_funder_name:
                    row = conn.execute(
                        text("EXEC FlaskHelperFunctions @Request='FunderIDDescription', @Text=:t"),
                        {"t": selected_funder_name},
                    ).fetchone()

                    if row:
                        funder_id = (
                            row[0]
                            if not hasattr(row, "_mapping")
                            else int(row._mapping.get("FunderID") or row[0])
                        )

                if not excel_report_type:
                    flash("Please choose an Excel export.", "warning")
                    return redirect(url_for("report_bp.new_reports"))

                if excel_report_type == "all_changes":
                    sql = text("""
                        SET NOCOUNT ON;
                        EXEC GetFunderYearGroupSummary_StudentWeighted_TY_LY_WithTrend
                            @CalendarYear = :year,
                            @Term = :term,
                            @NoChangeEpsilon = :change;
                    """)

                    df = pd.read_sql(
                        sql,
                        conn,
                        params={
                            "year": selected_year,
                            "term": selected_term,
                            "change":0.05
                        },
                    )

                    if df.empty:
                        flash("No data found for that export.", "warning")
                        return redirect(url_for("report_bp.new_reports"))

                    # Expected proc columns
                    required_cols = [
                        "Funder",
                        "YearGroupDesc",
                        "StudentCount_TY",
                        "StudentCount_LY",
                        "TY_YG_Rate",
                        "LY_YG_Rate",
                        "DeltaYG",
                        "TrendYG",
                    ]
                    missing_cols = [c for c in required_cols if c not in df.columns]
                    if missing_cols:
                        raise ValueError(
                            "The Excel export expected these columns but could not find them: "
                            + ", ".join(missing_cols)
                            + ". Actual columns were: "
                            + ", ".join(df.columns.astype(str).tolist())
                        )

                    filename = f"RatesByFunderYearGroup_T{selected_term}_{selected_year}.xlsx"

                    export_df = (
                        df[
                            [
                                "Funder",
                                "YearGroupDesc",
                                "StudentCount_TY",
                                "StudentCount_LY",
                                "TY_YG_Rate",
                                "LY_YG_Rate",
                                "DeltaYG",
                                "TrendYG",
                            ]
                        ]
                        .drop_duplicates(subset=["Funder", "YearGroupDesc"])
                        .copy()
                        .sort_values(["Funder", "YearGroupDesc"])
                    )

                    

                else:
                    flash("Unknown Excel export type.", "danger")
                    return redirect(url_for("report_bp.new_reports"))

            xlsx_id = uuid.uuid4().hex
            xlsx_path = REPORT_DIR / f"{xlsx_id}.xlsx"

            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                funder_col = "Funder"
                yeargroup_col = "YearGroupDesc"
                ty_count_col = "StudentCount_TY"
                ly_count_col = "StudentCount_LY"
                ty_col = "TY_YG_Rate"
                ly_col = "LY_YG_Rate"
                diff_col = "DeltaYG"
                trend_col = "TrendYG"

               
                overview_rows = [
                    ["Report", "Rates by Funder and Year Group"],
                    ["Calendar Year", selected_year],
                    ["Term", selected_term],
                    ["No-change threshold", "5 percentage points"],
                    ["", ""],
                    ["Sheet", "Description"],
                    ["Overview", "This sheet explains the workbook contents and definitions."],
                    ["Funder Summary", "One row per funder, with year groups as columns."],
                    ["All Funders", "Long-format table with all funder and year-group combinations."],
                    ["All Year Groups", "All funders for the combined year-group result."],
                    ["0-2", "All funders for year group 0-2."],
                    ["3-4", "All funders for year group 3-4."],
                    ["5-6", "All funders for year group 5-6."],
                    ["7-8", "All funders for year group 7-8."],
                    ["", ""],
                    ["Column", "Meaning"],
                    ["StudentCount_TY", "Student count for the selected year/term period."],
                    ["StudentCount_LY", "Student count for the comparison period."],
                    ["TY_YG_Rate", "Current period year-group rate."],
                    ["LY_YG_Rate", "Comparison period year-group rate."],
                    ["DeltaYG", "TY_YG_Rate minus LY_YG_Rate, shown in percentage points."],
                    ["TrendYG", "Direction of change: Up, Down, or No change."],
                ]

                overview_df = pd.DataFrame(overview_rows, columns=["Item", "Value"])
                overview_df.to_excel(writer, index=False, sheet_name="Overview")

                ws = writer.book["Overview"]

                overview_header_fill = PatternFill(
                    start_color="1A427D",
                    end_color="1A427D",
                    fill_type="solid"
                )
                overview_header_font = Font(color="FFFFFF", bold=True)
                header_rows = [1, 7, 17]
                for r in header_rows:
                    for cell in ws[r]:
                        cell.fill = overview_header_fill
                        cell.font = overview_header_font
                _autosize_and_freeze(ws)
                ws.freeze_panes = "A2"
                # =========================================
                # Middle sheets: one per year group
                # =========================================
                year_groups = export_df[yeargroup_col].dropna().unique().tolist()
                order = ["All Year Groups", "0-2", "3-4", "5-6", "7-8"]

                year_groups = sorted(year_groups, key=lambda x: order.index(x) if x in order else 999)

                for yg in year_groups:
                    yg_df = (
                        export_df[export_df[yeargroup_col] == yg]
                        .copy()
                        .sort_values(by=[funder_col])
                    )
                    sheet_name = _safe_sheet_name(str(yg), fallback="Year Group")
                    yg_df.to_excel(
                        writer,
                        index=False,
                        sheet_name=sheet_name,
                    )

                    ws = writer.book[sheet_name]
                    _convert_to_table(ws, f"YG_{sheet_name}")

                # =========================================
                # Final sheet: all funders (long)
                # =========================================
                all_df = export_df.copy().sort_values(by=[funder_col, yeargroup_col])
                all_df.to_excel(writer, index=False, sheet_name="All Funders")

                ws = writer.book["All Funders"]
                _convert_to_table(ws, "AllFunders")
                 # =========================================
                # Sheet 1: Funder Summary (wide)
                # =========================================
                summary_wide = export_df.pivot_table(
                    index=funder_col,
                    columns=yeargroup_col,
                    values=[ty_count_col, ly_count_col, ty_col, ly_col, diff_col],
                    aggfunc="first",
                )

                summary_wide.columns = [
                    f"{metric}_{group}" for metric, group in summary_wide.columns
                ]
                summary_wide = summary_wide.reset_index().sort_values(by=funder_col)

                summary_wide.to_excel(writer, index=False, sheet_name="Funder Summary")

                ws = writer.book["Funder Summary"]
                _convert_to_table(ws, "FunderSummary")
                # =========================================
                # Formatting
                # =========================================
                wb = writer.book

                # Header style
                header_fill = PatternFill(start_color="1a427d", end_color="1a427d", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)

                # Trend styles
                trend_up_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                trend_up_font = Font(color="FFFFFF", bold=True)

                trend_down_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
                trend_down_font = Font(color="FFFFFF", bold=True)

                trend_same_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                trend_same_font = Font(color="000000", bold=True)

                for ws in wb.worksheets:
                  
                    _autosize_and_freeze(ws)

                    header_map = {cell.column: cell.value for cell in ws[1]}

                    # Header row styling
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font

                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            header = header_map.get(cell.column)
                            if header is None:
                                continue

                            header_str = str(header)

                            # TY / LY rates as %
                            if (
                                header_str.startswith(ty_col)
                                or header_str.startswith(ly_col)
                                or header_str == ty_col
                                or header_str == ly_col
                            ):
                                if isinstance(cell.value, (int, float)) and cell.value is not None:
                                    cell.number_format = "0%"

                            # DeltaYG shown as percentage points
                            if header_str.startswith(diff_col) or header_str == diff_col:
                                if isinstance(cell.value, (int, float)) and cell.value is not None:
                                    cell.value = cell.value * 100
                                    cell.number_format = '+0.0;-0.0;0.0'

                            # Trend formatting
                            if header_str.startswith(trend_col) or header_str == trend_col:
                                val = str(cell.value).strip().lower()

                                if "up" in val:
                                    cell.fill = trend_up_fill
                                    cell.font = trend_up_font
                                elif "down" in val:
                                    cell.fill = trend_down_fill
                                    cell.font = trend_down_font
                                else:
                                    cell.fill = trend_same_fill
                                    cell.font = trend_same_font



            return send_file(
                xlsx_path,
                as_attachment=True,
                download_name=filename,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            current_app.logger.exception("❌ Excel export failed")
            try:
                log_alert(
                    email=session.get("user_email"),
                    role=session.get("user_role"),
                    entity_id=None,
                    link=request.url,
                    message=f"/Reports Excel export error: {str(e)}\n{traceback.format_exc()}"[:4000],
                )
            except Exception:
                pass

            flash("An error occurred while generating the Excel export.", "danger")
            return redirect(url_for("report_bp.new_reports"))


    if request.method == "POST" and action == "show_report":
        current_app.logger.info(
            "📩 /Reports POST show_report | ajax=%s role=%s user=%s",
            is_ajax,
            role,
            ctx["user"],
        )

        # initialise here so we can reference them in except logging
        funder_id = None
        region_id = None
        try:
            with engine.connect() as conn:
                current_app.logger.debug(
                    "📥 POST params | year=%s term=%s type=%s provider_id=%s school_id=%s",
                    selected_year,
                    selected_term,
                    selected_type,
                    selected_provider_id,
                    selected_school_id,
                )

                # Resolve funder from ADM dropdown (FUN is implied from session, PRO has none)
                if role == "ADM":
                    selected_funder_name = request.form.get("funder_name") or None
                    region_id = request.form.get("region_name") or None
                if selected_funder_name == "Loading funders…":
                    selected_funder_name = None

                current_app.logger.debug("effective funder_name=%s", selected_funder_name)
                current_app.logger.debug("effective region_name=%s", region_id)

                # If we have a funder *name*, resolve it to an ID
                if selected_funder_name:
                    row = conn.execute(
                        text("EXEC FlaskHelperFunctions @Request='FunderIDDescription', @Text=:t"),
                        {"t": selected_funder_name},
                    ).fetchone()

                    if not row:
                        msg = "Funder not found."

                        try:
                            log_alert(
                                email=session.get("user_email"),
                                role=session.get("user_role"),
                                entity_id=None,
                                link=request.url,
                                message=f"/Reports: funder not found for name '{selected_funder_name}'",
                            )
                        except Exception:
                            pass

                        if is_ajax:
                            return jsonify({"ok": False, "error": msg}), 400

                        flash(msg, "danger")
                        return redirect(url_for("report_bp.new_reports"))

                    funder_id = (
                        row[0]
                        if not hasattr(row, "_mapping")
                        else int(row._mapping.get("FunderID") or row[0])
                    )
                    current_app.logger.info("✅ resolved funder_id=%s", funder_id)

                # ---- Special case: funder_missing_data needs a funder *name* ----
                if selected_type == "funder_missing_data" and not selected_funder_name:
                    msg = "Please choose a funder for the missing data report."
                    try:
                        log_alert(
                            email=session.get("user_email"),
                            role=session.get("user_role"),
                            entity_id=None,
                            link=request.url,
                            message="/Reports: funder_missing_data selected but no funder chosen",
                        )
                    except Exception:
                        pass

                    if is_ajax:
                        return jsonify({"ok": False, "error": msg}), 400

                    flash(msg, "warning")
                    return render_template(
                        "reportingnew.html",
                        role=role,
                        funder_name=selected_funder_name,
                        results=None,
                        plot_payload=None,
                        plot_png_b64=None,
                        selected_term=selected_term,
                        selected_year=selected_year,
                        selected_type=selected_type,
                        entities_url=url_for("api_bp.get_entities"),
                        display=False,
                        no_data_banner=None,
                    )

                # ---- Generic validation for provider / funder / school ----
                validation_response = _validate_required_entities(
                    selected_type=selected_type,
                    role=role,
                    selected_provider_id=selected_provider_id,
                    selected_school_id=selected_school_id,
                    region_id = region_id,
                    funder_id=funder_id,
                    is_ajax=is_ajax,
                    selected_term=selected_term,
                    selected_year=selected_year,
                )
                if validation_response:
                    return validation_response

                fig = None

                current_app.logger.info("▶ executing report type=%s", selected_type)

                results, fig, no_data_banner_inner, early = _execute_report(
                    conn=conn,
                    selected_type=selected_type,
                    selected_year=selected_year,
                    selected_term=selected_term,
                    role=role,
                    region_id = region_id,
                    funder_id=funder_id,
                    selected_provider_id=selected_provider_id,
                    selected_school_id=selected_school_id,
                    selected_funder_name=selected_funder_name,
                    is_ajax=is_ajax,
                )
                if fig is not None and selected_type not in {"funder_weighted_average","provider_missing_classes", "funder_missing_classes", "national_competency_icons",}:

                    try:
                        footer_svg = os.path.join(current_app.static_folder, "footer.svg")
                        c = "#1a427d" if selected_type == "funder_missing_data" else "#1a427d40"
                        add_full_width_footer_svg(
                            fig,
                            footer_svg,
                            bottom_margin_frac=0.0,
                            max_footer_height_frac=0.20,
                            col_master=c,
                        )
                    except Exception as footer_e:
                        current_app.logger.info("⚠ Could not add footer: %s", footer_e)
                if early:
                    return early

                no_data_banner = no_data_banner_inner

                
                rows = results_to_rows(results)

                if(selected_type=="region_ly_vs_target"):
                    selected_term = 2
                    selected_year = 2025
                plot_payload = {
                    "year": selected_year,
                    "term": selected_term,
                    "type": selected_type,
                    "funder_id": funder_id,
                    "provider_id": int(selected_provider_id) if selected_provider_id else None,
                    "school_id": int(selected_school_id) if selected_school_id else None,
                    "rows": rows,   # ✅ JSON-safe always
                }

                if rows and fig is None:
                    fig, extra_banner = _build_figure_from_results(
                        selected_type,
                        results,  # keep original if some builders use it, but most use rows now
                        selected_term,
                        selected_year,
                        funder_id,
                        region_id,
                        selected_funder_name,
                        selected_provider_id,
                        selected_school_id,
                        rows,
                    )
                    if extra_banner:
                        no_data_banner = extra_banner
                
                if selected_type in {"funder_student_count", "funder_progress_summary", "funder_teacher_review_summary", "provider_missing_classes","funder_missing_classes",  "national_competency_icons",  "region_coverage_report",}:
                    if fig is not None:
                        PREFIX_MAP = {
                            "funder_student_count": "Funder_Student_Count",
                            "funder_progress_summary": "Funder_Progress_Summary",
                            "funder_teacher_review_summary": "Funder_Teacher_Reviews_Summary",
                            "provider_missing_classes": "provider_missing_classes",
                            "funder_missing_classes": "funder_missing_classes",
                            "region_coverage_report": "region_coverage_report",
                            "national_competency_icons": "National_Competency_Icons",
                        }

                        prefix = PREFIX_MAP.get(selected_type, "Report")

                        provider_label = (request.form.get("provider_name") or "").strip() or f"Provider_{selected_provider_id or ''}"

                        name_for_file = (
                            provider_label
                            if selected_type == "provider_missing_classes"
                            else region_id
                            if selected_type == "region_coverage_report"
                            else "National"
                            if selected_type == "national_competency_icons"
                            else selected_funder_name
                        )

                        plot_png_b64 = _persist_preview_for_existing_report(
                            report_id=session["report_id"],
                            fig=fig,
                            selected_term=selected_term,
                            selected_year=selected_year,
                            selected_funder_name=name_for_file,
                            base_label_prefix=prefix,
                        )
                        display = True
                        params_json = json.dumps({
                            "report_type": selected_type,
                            "calendar_year": selected_year,
                            "term": selected_term,
                            "region": region_id,
                            "funder_id": funder_id,
                            "provider_id": selected_provider_id,
                            "school_id": selected_school_id,
                            "excel_report_option": excel_report_type,
                        }, default=str)
                        log_report_run(
                            engine=engine,
                            report_name=selected_type,
                            year=selected_year,
                            term=selected_term,
                            region_name=region_id,
                            funder_id=funder_id,
                            provider_id=selected_provider_id,
                            school_id=selected_school_id,
                            params_json=params_json,
                        )
                
                else:
                    # Existing behaviour for all other report types
                    if fig is not None:
                        plot_png_b64 = _persist_figure_and_session(
                            fig,
                            selected_type,
                            selected_term,
                            selected_year,
                            selected_funder_name,
                            selected_provider_id,
                            selected_school_id,
                            results,
                        )
                        display = True
                        params_json = json.dumps({
                            "report_type": selected_type,
                            "calendar_year": selected_year,
                            "term": selected_term,
                            "region": region_id,
                            "funder_id": funder_id,
                            "provider_id": selected_provider_id,
                            "school_id": selected_school_id,
                            "excel_report_option": excel_report_type,
                        }, default=str)
                        log_report_run(
                            engine=engine,
                            report_name=selected_type,
                            year=selected_year,
                            term=selected_term,
                            region_name=region_id,
                            funder_id=funder_id,
                            provider_id=selected_provider_id,
                            school_id=selected_school_id,
                            params_json=params_json,
                        )

                # ===== AJAX response (no full reload) =====
                if is_ajax:
                    provider_name = request.form.get("provider_name")
                    school_name = request.form.get("school_name")
                    provider_id = request.form.get("provider_id")
                    school_id = request.form.get("school_id")

                    left_bits = []
                    if selected_funder_name:
                        left_bits.append(selected_funder_name)
                    left_bits.append(f"Term {selected_term}, {selected_year}")

                    if selected_type == "school_ytd_vs_target" and (school_name or school_id):
                        left_bits.append(school_name or f"School MOE {school_id}")
                    elif provider_name:
                        left_bits.append(provider_name)
                    elif provider_id:
                        left_bits.append(f"Provider ID {provider_id}")

                    header_html = " • ".join(left_bits)
                    allow_png = bool(display) and (
                        selected_type not in {
                            "funder_student_count",
                            "funder_progress_summary",
                            "funder_teacher_review_summary",
                            "provider_missing_classes",
                               "funder_missing_classes", 
                        }
                    )
                    return jsonify(
                        {
                            "ok": True,
                            "plot_png_b64": plot_png_b64,
                            "header_html": header_html,
                            "display": bool(display),
                            "allow_png": bool(allow_png),
                            "notice": no_data_banner,
                        }
                    )


        except Exception as e:
            # one clean traceback in logs (no print spam)
            current_app.logger.exception(
                "❌ /Reports failed | role=%s year=%s term=%s type=%s funder_name=%s funder_id=%s provider_id=%s school_id=%s user=%s ajax=%s",
                role,
                selected_year,
                selected_term,
                selected_type,
                selected_funder_name,
                funder_id,
                selected_provider_id,
                selected_school_id,
                session.get("user_email"),
                is_ajax,
            )
            params_json = json.dumps({
                "report_type": selected_type,
                "calendar_year": selected_year,
                "term": selected_term,
                "region": region_id,
                "funder_id": funder_id,
                "provider_id": selected_provider_id,
                "school_id": selected_school_id,
                "excel_report_option": excel_report_type,
            }, default=str)
            log_report_run(
                engine=engine,
                report_name=selected_type,
                year=selected_year,
                term=selected_term,
                region_name=region_id,
                funder_id=funder_id,
                provider_id=selected_provider_id,
                school_id=selected_school_id,
                params_json=params_json,
                success=False,
                error_message=str(e),
            )
            err_text = str(getattr(e, "orig", e))
            tb = traceback.format_exc()

            # ✅ log to AUD_Alerts (best-effort)
            try:
                log_alert(
                    email=session.get("user_email"),
                    role=session.get("user_role"),
                    entity_id=None,
                    link=request.url,
                    message=f"/Reports error: {err_text}\n{tb}"[:4000],
                )
            except Exception:
                pass

            # Friendly message + status code (unchanged behaviour)
            if "Provider is not linked to the supplied FunderID" in err_text:
                user_msg = "You must select a provider and funder that are linked for this report"
                status_code = 400
            else:
                user_msg = "An error occurred while generating the report."
                status_code = 500

            if is_ajax:
                return jsonify({"ok": False, "error": user_msg, "display": False}), status_code

            flash(user_msg, "danger")
            return redirect(url_for("report_bp.new_reports"))

    # Post-sort for stable ordering
    if results and isinstance(results, list) and results and "CompetencyDesc" in results[0]:
        results = sorted(results, key=lambda x: (x.get("CompetencyDesc") or ""))

    # GET, or POST without show_report → just render the page; dropdowns loaded via AJAX
    return render_template(
        "reportingnew.html",
        role=role,
        funder_name=selected_funder_name,
        results=results,
        plot_payload=plot_payload,
        plot_png_b64=plot_png_b64,
        selected_term=selected_term,
        selected_year=selected_year,
        selected_type=selected_type,
        entities_url=url_for("api_bp.get_entities"),
        display=display,
        no_data_banner=no_data_banner,
    )