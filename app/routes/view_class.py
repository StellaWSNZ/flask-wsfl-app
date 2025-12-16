# app/routes/view_class.py

# ---------------------------
# Standard library
# ---------------------------
import base64
import io
import json
import re
import traceback
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from urllib.parse import urlencode
import qrcode
import csv
import xlrd

# ---------------------------
# Third-party
# ---------------------------
import matplotlib
matplotlib.use("Agg")  # Prevent GUI backend errors on servers
import pandas as pd
import pyodbc
from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.exc import DBAPIError, SQLAlchemyError


# ---------------------------
# Local imports
# ---------------------------
from app.routes.auth import login_required
from app.utils.database import get_db_engine, get_terms, get_years, log_alert

# ---------------------------
# Blueprint
# ---------------------------
class_bp = Blueprint("class_bp", __name__)

# =========================
# Tiny Helpers
# =========================

def _require_int(v, name):
    try:
        return int(v)
    except (TypeError, ValueError):
        raise ValueError(f"Invalid {name}")

def _safe_filename(name: str) -> str:
    name = re.sub(r'[\\/*?:"<>|]+', "-", str(name))
    name = re.sub(r"\s+", " ", name).strip()
    return name[:200]

def friendly_sql_error(exc: Exception) -> tuple[int, str, int | None]:
    """
    Returns (http_status, user_message, sql_error_number) for a DB exception.
    Works with pyodbc/SQL Server via SQLAlchemy.
    """
    raw = str(getattr(exc, "orig", exc))

    # Best-effort extract of SQL Server native error number:
    # e.g. "... (50010) (SQLExecDirectW)"
    m = re.search(r"\((\d{5,7})\)\s*\(SQL", raw)  # preferred
    code = int(m.group(1)) if m else None
    if code is None:
        # Fallback: grab a 4–7 digit number if present anywhere
        m2 = re.search(r"\b(\d{4,7})\b", raw)
        code = int(m2.group(1)) if m2 else None

    MAP = {
        50001: (400, "NSN must be numeric."),
        50002: (400, "First name and last name are required."),
        50003: (400, "That class could not be found."),
        50004: (400, "That class is missing its year/term setup."),
        50006: (500, "We couldn't save the student record. Please try again."),
        50007: (500, "We couldn't add the student to the class. Please try again."),
        50008: (409, "Another user updated this student at the same time. Please try again."),
        50010: (409, "That NSN already exists. Use Search → Add to put them into this class."),
    }
    if code in MAP:
        http, msg = MAP[code]
        return http, msg, code

    # Common SQL Server duplicate key
    if code in (2627, 2601):
        return 409, "A record with this key already exists.", code

    # Uncommittable transaction (3930) message text
    if "cannot be committed and cannot support operations" in raw:
        return 500, "We hit a database error and rolled back your changes. Please try again.", code

    # Fallback
    return 500, "Something went wrong saving this student. Please try again.", code

def _json_error(msg, code=400):
    return jsonify({"ok": False, "error": msg}), code

# =========================
# Role / device / request helpers
# =========================
def _require_moe_or_adm():
    role = session.get("user_role")
    admin = int(session.get("user_admin") or 0)
    return (role == "MOE" and admin == 1)or role == "ADM"

def _require_moe_or_adm2():
    role = session.get("user_role")
    admin = int(session.get("user_admin") or 0)
    return (role == "MOE")or role == "ADM"

def is_mobile() -> bool:
    """Very simple UA sniff to block phones/tablets."""
    try:
        ua = (request.headers.get("User-Agent") or "").lower()
    except Exception as e:
        current_app.logger.debug("UA read failed: %r", e)
        return False
    return any(k in ua for k in ("iphone", "android", "ipad", "mobile"))


# =========================
# DB “service” helper
# =========================
def _ensure_authorised_for_class(engine, class_id: int) -> None:
    if session.get("user_role") is None:
        abort(403, description="You are not authorised to view that page.")

def _get_class_meta(engine, class_id: int):
    # Try proc (preferred)
    try:
        with engine.begin() as conn:
            row = conn.execute(
                text("EXEC FlaskGetClassMeta @ClassID = :cid"),
                {"cid": class_id}
            ).mappings().first()
        if row:
            return {
                "ClassName": row.get("ClassName") or f"Class {class_id}",
                "TeacherName": row.get("TeacherName") or "",
                "SchoolName": row.get("SchoolName") or "",
                "MOENumber": row.get("MOENumber"),
            }
    except SQLAlchemyError:
        pass
    except Exception:
        pass

    return {"ClassName": f"Class {class_id}", "TeacherName": "", "SchoolName": "", "MOENumber": None}

def _load_class_list_df(engine, class_id: int, term: int, year: int) -> pd.DataFrame:
    """Replace this EXEC with your real exporter for class list."""
    current_app.logger.info(session.get("user_role"))
    with engine.begin() as conn:
        df = pd.read_sql(
            text("EXEC FlaskExportClassList @ClassID=:cid, @Term=:t, @CalendarYear=:y, @Role=:r")
,
            conn, params={"cid": class_id, "t": term, "y": year, "r":session.get("user_role")}
        )
    # Optional: preferred ordering
    #lead = [c for c in ["NSN","LastName","FirstName","PreferredName","YearLevelID","DateOfBirth"] if c in df.columns]
    #rest = [c for c in df.columns if c not in lead]
    return  df

def _load_achievements_df(engine, class_id: int, term: int, year: int) -> pd.DataFrame:
    """Replace this EXEC with your real exporter for achievements table (one row per student)."""
    with engine.begin() as conn:
        df = pd.read_sql(
            text("EXEC FlaskExportAchievements @ClassID=:cid, @Term=:t, @Year=:y"),
            conn, params={"cid": class_id, "t": term, "y": year}
        )
    # Bring identity columns to the front if present
    lead = [c for c in ["NSN","LastName","PreferredName","YearLevelID"] if c in df.columns]
    rest = [c for c in df.columns if c not in lead]
    return df[lead + rest] if lead else df


# =========================
# Export/format helpers
# =========================
def generate_qr_code_png(data, box_size=2):
    qr = qrcode.QRCode(box_size=box_size, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    
    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    img_str = base64.b64encode(buffered.getvalue()).decode("utf-8")
    return f"data:image/png;base64,{img_str}"

def excel_bytes_writer(df: pd.DataFrame, sheet_name: str = "Sheet1"):
    """
    Writes a compact, readable Excel:
    - Wrapped headers (supports \n in header text)
    - Narrow default widths (12), slightly wider for name columns
    - Centered numbers/booleans, wrapped text for others
    """
    bio = io.BytesIO()
    sheet = (sheet_name or "Sheet1")[:31]

    try:
        with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet)

            wb = writer.book
            ws = writer.sheets[sheet]

            # Formats
            header_fmt = wb.add_format({
                "bold": True, "valign": "top", "text_wrap": True,
                "border": 1, "bg_color": "#F2F2F2"
            })
            text_fmt   = wb.add_format({"valign": "top", "text_wrap": True})
            num_fmt    = wb.add_format({"valign": "vcenter", "align": "center"})

            # Re-write headers with wrapping (supports \n inserted earlier)
            for j, col in enumerate(df.columns):
                ws.write(0, j, str(col), header_fmt)

            # Make header row a bit taller for wraps
            ws.set_row(0, 32)

            # Column width plan
            default_width = 12
            width_map = {
                "NSN": 8,
                "YearLevelID": 8,
                "LastName": 16,
                "Surname": 16,
                "FirstName": 14,
                "PreferredName": 14,
                "DateOfBirth": 11,
            }

            # Apply widths + sensible default cell formats
            for j, col in enumerate(df.columns):
                col_name = str(col)
                width = width_map.get(col_name, default_width)

                # Choose a default format for the column
                series = df[col]
                if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_bool_dtype(series):
                    col_fmt = num_fmt
                else:
                    col_fmt = text_fmt

                ws.set_column(j, j, width, col_fmt)

            # Freeze header
            ws.freeze_panes(1, 0)

        bio.seek(0)
        return bio

    except Exception:
        # Fallback (no styling) if xlsxwriter is missing
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet)
            # Optional: set simple widths in openpyxl
            try:
                from openpyxl.utils import get_column_letter
                ws = writer.sheets[sheet]
                for j, col in enumerate(df.columns, start=1):
                    col_name = str(col)
                    width = width_map.get(col_name, default_width)
                    ws.column_dimensions[get_column_letter(j)].width = width
            except Exception:
                pass
        bio.seek(0)
        return bio

# =========================
# Print/ PDF Pipeline Helpers
# =========================
def _render_print_html(engine,moe_number: int, class_id: int, term: int, year: int, filter_type: str, order_by: str) -> str:
    # Reuse your existing context builder; you already had this idea earlier
    # If you don't have a _build_print_context yet, we can synthesize a tiny wrapper
    ctx = _build_print_context(engine,moe_number,  class_id, term, year, filter_type, order_by)
    return render_template("print_view.html", **ctx)

def _html_to_pdf_bytes_with_playwright(html: str, base_url: str | None = None) -> bytes | None:
    """
    Returns PDF bytes using Playwright/Chromium, or None if Playwright isn't available
    or PDF rendering fails for any reason.
    """
    if not sync_playwright:
        return None

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--no-sandbox"])
            context = browser.new_context()
            page = context.new_page()
            # set_content supports base_url so relative assets resolve
            page.set_content(html, base_url=base_url, wait_until="load")
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                margin={"top": "10mm", "right": "10mm", "bottom": "10mm", "left": "10mm"},
            )
            browser.close()
            return pdf_bytes
    except Exception:
        current_app.logger.exception("Playwright PDF generation failed")
        return None

def _build_print_context(engine,moe_number:int, class_id: int, term: int, year: int, filter_type: str, order_by: str):
    """
    Build the same context dict that print_class_view uses to render print_view.html.
    Reuses cache when possible; regenerates if needed.
    """
    cache_key = f"{class_id}_{term}_{year}_{filter_type}"
    class_cache = session.get("class_cache", {})
    cache = class_cache.get(cache_key)

    # If missing, rebuild like print_class_view
    if not cache or "student_competencies" not in cache:
        with engine.begin() as conn:
            result = conn.execute(
                text("""EXEC FlaskGetClassStudentAchievement 
                        @ClassID = :class_id, 
                        @Term = :term, 
                        @CalendarYear = :year, 
                        @Email = :email, 
                        @FilterType = :filter"""),
                {"class_id": class_id, "term": term, "year": year,
                 "email": session.get("user_email"), "filter": filter_type}
            )
            df = pd.DataFrame(result.fetchall(), columns=result.keys())
            if df.empty:
                # Return a minimal context; caller can handle “no data”
                return {
                    "grouped": {"0–2": [], "3–4": [], "5–6": [], "7–8": []},
                    "columns_by_range": {"0–2": [], "3–4": [], "5–6": [], "7–8": []},
                    "class_name": "(Unknown)", "teacher_name": "(Unknown)",
                    "filter_type": filter_type, "now": datetime.now,
                    "qr_data_uri": generate_qr_code_png(url_for("auth_bp.login", _external=True))
                }

            comp_df = (
                df[["CompetencyLabel", "CompetencyID", "YearGroupID"]]
                .drop_duplicates()
                .rename(columns={"CompetencyLabel": "label"})
            )
            comp_df["col_order"] = comp_df["YearGroupID"].astype(str).str.zfill(2) + "-" + comp_df["CompetencyID"].astype(str).str.zfill(4)
            comp_df = comp_df.sort_values("col_order")

            meta_cols = [
                "NSN", "FirstName", "LastName", "PreferredName",
                "DateOfBirth", "Ethnicity", "YearLevelID"
            ]
            df_combined = df.pivot_table(
                index=meta_cols,
                columns="label",
                values="CompetencyStatus",
                aggfunc="first"
            ).fillna(0).astype(int).replace({1: "✓", 0: ""}).reset_index()

            expiry_time = datetime.now(timezone.utc) + timedelta(minutes=15)
            class_cache[cache_key] = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "expires": expiry_time.isoformat(),
                "students": df_combined.to_dict(),
                "competencies": comp_df.to_dict(orient="records"),
                "filter": filter_type,
                "student_competencies": df_combined.to_dict(orient="records"),
            }
            session["class_cache"] = class_cache

    else:
        df_combined = pd.DataFrame(cache["student_competencies"]).replace({1: "✓", 0: ""})
        comp_df = pd.DataFrame(cache["competencies"])

    # Column groups for the template
    labels = comp_df["label"].tolist() if not comp_df.empty else []
    def _labels_for(yr): return [l for l in labels if f"({yr})" in l]
    columns_by_range = {
        "0–2": _labels_for("0-2"),
        "3–4": _labels_for("3-4"),
        "5–6": _labels_for("5-6"),
        "7–8": _labels_for("7-8"),
    }

    # Grouped rows (template expects same rows per group; columns are filtered per range)
    grouped = {"0–2": [], "3–4": [], "5–6": [], "7–8": []}
    for row in df_combined.to_dict(orient="records"):
        for k in grouped.keys():
            grouped[k].append(row)

    # Class/teacher names
    with engine.connect() as conn:
        class_info = conn.execute(
            text("EXEC FlaskHelperFunctions @Request = :Request, @Number = :class_id"),
            {"Request": "ClassInfoByID", "class_id": class_id}
        ).fetchone()
    class_name   = class_info.ClassName if class_info else "Unknown Class"
    teacher_name = class_info.TeacherName if class_info else "Unknown Teacher"

    # QR for login-to-view
    target_path = url_for("class_bp.view_class",moe_number = moe_number,  class_id=class_id, term=term, year=year)
    login_url   = url_for("auth_bp.login", next=target_path, _external=True)
    qr_data_uri = generate_qr_code_png(login_url)

    # Sort order for display (optional)
    key_col = "PreferredName" if order_by == "first" else "LastName"
    if grouped["0–2"] and key_col in grouped["0–2"][0]:
        for k in grouped.keys():
            grouped[k] = sorted(grouped[k], key=lambda r: (r.get(key_col) or "").lower())

    return {
        "grouped": grouped,
        "columns_by_range": columns_by_range,
        "class_name": class_name,
        "teacher_name": teacher_name,
        "filter_type": filter_type,
        "now": datetime.now,
        "qr_data_uri": qr_data_uri,
    }
def _allowed_school_ids(conn, include_inactive: int = 0) -> set[int]:
    user_id   = session.get("user_id")
    user_role = session.get("user_role")

    try:
        user_id = int(user_id) if user_id is not None else None
    except (TypeError, ValueError):
        user_id = None

    rows = conn.execute(
        text("""
            SET NOCOUNT ON;
            EXEC dbo.FlaskGetEntities
                @EntityType      = :EntityType,
                @Role            = :Role,
                @ID              = :ID,
                @IncludeInactive = :IncludeInactive;
        """),
        {
            "EntityType": "School",
            "Role": user_role,
            "ID": user_id,
            "IncludeInactive": int(include_inactive or 0),
        },
    ).mappings().all()

    # Assumes SP returns ID = MOENumber for schools
    return {int(r["ID"]) for r in rows if r.get("ID") is not None}

# =========================
# Page Routes
# =========================
@class_bp.route('/Classes/<int:moe_number>/<int:class_id>/<int:term>/<int:year>')
@login_required
def view_class(moe_number,class_id, term, year):
    try:
        # ---------- Query params ----------
        filter_type = request.args.get("filter", "all")
        order_by    = request.args.get("order_by", "last")

        # ---------- Cache lookup ----------
        cache_key   = f"{class_id}_{term}_{year}_{filter_type}"
        class_cache = session.get("class_cache", {})
        cached      = class_cache.get(cache_key)
        
        engine = get_db_engine()
        with engine.begin() as conn:

            # -----------------------------
            # Authorisation (school access)
            # -----------------------------
            role = (session.get("user_role") or "").upper()

            if role == "MOE":
                session_moe = session.get("moe_number") or session.get("user_id") or session.get("ID")
                try:
                    session_moe = int(session_moe)
                except (TypeError, ValueError):
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403

                if int(moe_number) != session_moe:
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403

            else:
                allowed = _allowed_school_ids(conn, include_inactive=0)
                if int(moe_number) not in allowed:
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403

        # ❌ No valid cache → fetch from DB
        engine = get_db_engine()
        with engine.begin() as conn:
            # Scenarios
            scenario_result = conn.execute(
                text("EXEC FlaskHelperFunctions @Request = :request"),
                {"request": "Scenario"}
            )
            scenarios = [dict(row._mapping) for row in scenario_result]

           
            result = conn.execute(
                text("""
                    EXEC FlaskGetClassStudentAchievement 
                        @ClassID = :class_id, 
                        @Term = :term, 
                        @CalendarYear = :year, 
                        @Email = :email, 
                        @FilterType = :filter
                """),
                {
                    "class_id": class_id,
                    "term": term,
                    "year": year,
                    "email": session.get("user_email"),
                    "filter": filter_type
                }
            )
            rows = result.fetchall()
            if not rows:
                current_app.logger.info("ℹ️ No rows returned for this class/term/year/filter.")
                return render_template(
                    "student_achievement.html",
                    students=[],
                    columns=[],
                    competency_id_map={},
                    scenarios=scenarios,
                    class_id=class_id,
                    class_name="(Unknown)",
                    teacher_name="(Unknown)",
                    school_name="(Unknown)",
                    class_title="No data for this selection",
                    edit=session.get("user_admin"),
                    autofill_map={},
                    term=term,
                    year=year,
                    moe_number = moe_number,
                    order_by=order_by,
                    filter_type=filter_type
                )

            df = pd.DataFrame(rows, columns=result.keys())

            # Build comp_df for ordering/map BEFORE dropping cols
            need_cols = ["CompetencyLabel", "CompetencyID", "YearGroupID"]
            have_cols = [c for c in need_cols if c in df.columns]
            comp_df = (
                df[have_cols]
                .drop_duplicates()
                .rename(columns={"CompetencyLabel": "label"})
            ) if have_cols else pd.DataFrame(columns=["label","CompetencyID","YearGroupID"])

            # Normalize labels to avoid invisible mismatches
            if "label" in comp_df.columns:
                comp_df["label"] = comp_df["label"].astype(str).str.strip()

            # Titles (guard if columns missing)
            class_name   = df["ClassName"].dropna().unique()[0]   if "ClassName"   in df.columns and df["ClassName"].notna().any()   else "(Unknown)"
            teacher_name = df["TeacherName"].dropna().unique()[0] if "TeacherName" in df.columns and df["TeacherName"].notna().any() else "(Unknown)"
            school_name  = df["SchoolName"].dropna().unique()[0]  if "SchoolName"  in df.columns and df["SchoolName"].notna().any()  else "(Unknown)"
            title_string = f"Class Name: {class_name} | Teacher Name: {teacher_name} | School Name: {school_name}"

            # Drop meta columns we don't want duplicated post-pivot
            drop_cols = [c for c in ["ClassName", "TeacherName", "SchoolName", "CompetencyID", "YearGroupID"] if c in df.columns]
            df = df.drop(columns=drop_cols)

            # Pivot
            meta_cols = [
                "NSN", "FirstName", "LastName", "PreferredName",
                "Ethnicity", "YearLevelID", "Scenario1", "Scenario2"
            ]
            existing_meta = [c for c in meta_cols if c in df.columns]
            pivot_df = df.pivot_table(
                index=existing_meta,
                columns="CompetencyLabel",
                values="CompetencyStatus",
                aggfunc="first"
            ).reset_index()

            # Normalize pivot column labels too
            pivot_df.columns = [str(c).strip() for c in pivot_df.columns]

            # Desired competency order
            if not comp_df.empty:
                comp_df_sorted = comp_df.sort_values(["YearGroupID", "CompetencyID"])
            else:
                comp_df_sorted = pd.DataFrame(columns=["label","CompetencyID","YearGroupID"])

            desired_competencies = comp_df_sorted["label"].tolist()

            # Sort students by requested key if present
            key_col = "PreferredName" if order_by == "first" else "LastName"
            if key_col in pivot_df.columns:
                pivot_df = pivot_df.sort_values(
                    by=key_col,
                    key=lambda col: col.fillna("").astype(str).str.lower()

                )

            # Rename scenario columns (only if present)
            rename_map = {
                "Scenario1": "Scenario One - Selected <br>(7-8)",
                "Scenario2": "Scenario Two - Selected <br>(7-8)"
            }
            rename_applied = {k: v for k, v in rename_map.items() if k in pivot_df.columns}
            if rename_applied:
                pivot_df = pivot_df.rename(columns=rename_applied)

            # Fixed & scenario columns
            existing_cols = set(pivot_df.columns)
            fixed_cols_all = ["NSN", "LastName", "PreferredName", "YearLevelID"]
            fixed_cols_present = [c for c in fixed_cols_all if c in existing_cols]

            scenario_cols_all = [
                "Scenario One - Selected <br>(7-8)",
                "Scenario One - Completed <br>(7-8)",
                "Scenario Two - Selected <br>(7-8)",
                "Scenario Two - Completed <br>(7-8)"
            ]
            existing_scenario_cols = [c for c in scenario_cols_all if c in existing_cols]
            scenario_set = set(existing_scenario_cols)

            # ===== Force-include all competencies (even if no rows) =====
            # Exclude any labels that equal scenario headers
            full_comp_cols = [lbl for lbl in desired_competencies if lbl not in scenario_set]

            forced_added = []
            for lbl in full_comp_cols:
                if lbl not in pivot_df.columns:
                    pivot_df[lbl] = pd.NA
                    forced_added.append(lbl)
            if forced_added:
                current_app.logger.info("ℹ️ Forced in empty competency columns (no rows under current filter).")

            # Final column order (only columns that exist + forced)
            ordered_cols = fixed_cols_present + full_comp_cols + existing_scenario_cols
            ordered_cols = [c for c in ordered_cols if c in pivot_df.columns]  # safety
            pivot_df = pivot_df[ordered_cols]

            # Build competency_id_map for template
            competency_id_map = {}
            if not comp_df_sorted.empty and {"label","CompetencyID"} <= set(comp_df_sorted.columns):
                competency_id_map = comp_df_sorted.set_index("label")["CompetencyID"].to_dict()

            # Autofill map
            auto_result = conn.execute(
                text("EXEC FlaskHelperFunctions @Request = :request"),
                {"request": "AutoMappedCompetencies"}
            )
            header_map = defaultdict(list)
            for row in auto_result:
                header_map[row.HeaderPre].append(row.HeaderPost)

            # Cache it
            expiry_time = datetime.now(timezone.utc) + timedelta(minutes=15)
            class_cache[cache_key] = {
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "expires": expiry_time.isoformat(),
                "students": pivot_df.to_dict(),  # raw table
                "competencies": comp_df_sorted.to_dict(orient="records"),
                "filter": filter_type,
                "student_competencies": pivot_df.to_dict(orient="records"),  # for cached branch
                "class_name": class_name,
                "teacher_name": teacher_name,
                "school_name": school_name,
                "scenarios": scenarios,
                "autofill_map": dict(header_map)
            }
            session["class_cache"] = class_cache
            target = "Basic awareness of potential water-related hazards"
            cols_list = list(pivot_df.columns)

            # Also log what the template will receive:
            render_cols = [c for c in pivot_df.columns if c not in ["DateOfBirth","Ethnicity","FirstName","NSN"]]
            
            # Render
            return render_template(
                "student_achievement.html",
                students=pivot_df.to_dict(orient="records"),
                columns=[c for c in pivot_df.columns if c not in ["DateOfBirth", "Ethnicity", "FirstName", "NSN"]],
                competency_id_map=competency_id_map,
                scenarios=scenarios,
                class_id=class_id,
                class_name=class_name,
                teacher_name=teacher_name,
                school_name=school_name,
                class_title=title_string,
                edit=session.get("user_admin"),
                autofill_map=header_map,
                moe_number = moe_number,
                term=term,
                year=year,
                order_by=order_by,
                filter_type=filter_type
            )

    except Exception as e:
        # console trace for local dev
        traceback.print_exc()

        # structured app log
        try:
            current_app.logger.error(
                "❌ view_class crashed | class_id=%s term=%s year=%s | user=%s | filter=%s | order_by=%s",
                class_id,
                term,
                year,
                session.get("user_email"),
                request.args.get("filter", "all"),
                request.args.get("order_by", "last"),
                exc_info=True
            )
        except Exception:
            pass  # logging must never break

        # best-effort alert in DB (won't raise)
        try:
            log_alert(
                email    = session.get("user_email"),
                role     = session.get("user_role"),
                entity_id= session.get("user_id"),
                link     = url_for("class_bp.view_class", class_id=class_id, term=term, year=year, _external=True),
                message  = f"/Class/{class_id}/{term}/{year} failed (filter={request.args.get('filter')}, order_by={request.args.get('order_by')}): {e}\n{traceback.format_exc()}"[:4000],
            )
        except Exception:
            pass

        # user-facing response unchanged
        return "An internal error occurred. Check logs for details.", 500


def _entities_for_user(conn, entity_type: str, include_inactive: int = 0) -> list[dict]:
    user_id   = session.get("user_id")
    user_role = session.get("user_role")

    try:
        user_id = int(user_id) if user_id is not None else None
    except (TypeError, ValueError):
        user_id = None

    sql = text("""
        SET NOCOUNT ON;
        EXEC dbo.FlaskGetEntities
            @EntityType      = :EntityType,
            @Role            = :Role,
            @ID              = :ID,
            @IncludeInactive = :IncludeInactive;
    """)

    rows = conn.execute(sql, {
        "EntityType": entity_type,
        "Role": user_role,
        "ID": user_id,
        "IncludeInactive": int(include_inactive or 0),
    }).mappings().all()

    return [{"id": int(r["ID"]), "description": str(r["Description"])} for r in rows]

@class_bp.route("/FilterClasses", methods=["GET", "POST"])
@login_required
def filter_classes():

    user_role  = (session.get("user_role") or "").upper()
    user_admin = int(session.get("user_admin", 0))

    terms = get_terms()
    years = get_years()

    classes = []
    suggestions = []
    schools = []          # IMPORTANT: empty → AJAX fills it
    moe_number = None

    # selections (safe casting)
    selected_term = request.form.get("term") or session.get("nearest_term")
    selected_year = request.form.get("calendaryear") or session.get("nearest_year")
    selected_moe  = request.form.get("moe_number")

    # MOE school inference
    if user_role == "MOE":
        moe_number = session.get("moe_number") or session.get("user_id")

    if request.method == "POST":
        try:
            term = int(selected_term)
            year = int(selected_year)
            if term not in (1, 2, 3, 4):
                raise ValueError("Invalid term")

            if user_role == "MOE":
                if not moe_number:
                    raise ValueError("Could not determine your school")
                target_moe = int(moe_number)
            else:
                if not selected_moe:
                    raise ValueError("Please select a school")
                target_moe = int(selected_moe)

            engine = get_db_engine()
            with engine.begin() as conn:
                rows = conn.execute(
                    text("""
                        EXEC FlaskHelperFunctionsSpecific
                            @Request      = 'ClassesBySchoolTermYear',
                            @MOENumber    = :moe,
                            @Term         = :term,
                            @CalendarYear = :year
                    """),
                    {"moe": target_moe, "term": term, "year": year}
                ).fetchall()

                classes = [dict(r._mapping) for r in rows]

                if not classes:
                    sugg = conn.execute(
                        text("""
                            EXEC FlaskHelperFunctionsSpecific
                                @Request   = 'DistinctTermsForSchool',
                                @MOENumber = :moe
                        """),
                        {"moe": target_moe}
                    ).fetchall()
                    suggestions = [r.Label for r in sugg if r.Label]

            if user_role != "MOE":
                moe_number = target_moe

        except Exception:
            current_app.logger.exception("Error loading classes")
            flash("Could not load classes.", "danger")

    return render_template(
        "classes.html",
        user_role=user_role,
        user_admin=user_admin,
        terms=terms,
        years=years,
        classes=classes,
        suggestions=suggestions,
        schools=schools,          # always empty
        moe_number=moe_number,
        selected_term=selected_term,
        selected_year=selected_year,
        desc = session.get("desc"),
    )


@class_bp.route("/Class/print/<int:moe_number>/<int:class_id>/<int:term>/<int:year>")
@login_required
def print_class_view(moe_number, class_id, term, year):
    try:
        engine = get_db_engine()
        with engine.begin() as conn:

            # -----------------------------
            # Authorisation (school access)
            # -----------------------------
            role = (session.get("user_role") or "").upper()

            if role == "MOE":
                session_moe = session.get("moe_number") or session.get("user_id") or session.get("ID")
                try:
                    session_moe = int(session_moe)
                except (TypeError, ValueError):
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403

                if int(moe_number) != session_moe:
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403

            else:
                allowed = _allowed_school_ids(conn, include_inactive=0)
                if int(moe_number) not in allowed:
                    return render_template(
                "error.html",
                error="You are not authorised to view that page.",
                code=403
            ), 403
        filter_type = request.args.get("filter") or session.get("last_filter_used", "all")
        order_by    = request.args.get("order_by", "last")

        engine = get_db_engine()
        ctx = _build_print_context(engine, moe_number, class_id, term, year, filter_type, order_by)

        # If no data, you can redirect or render a minimal page:
        return render_template("print_view.html", **ctx)

    except Exception as e:
        current_app.logger.exception("❌ Unhandled error in print_class_view.")
        return "Internal Server Error (print view)", 500

@class_bp.route("/EditClass")
@login_required
def class_students_page():
    # role check
    if not _require_moe_or_adm2():
        return _json_error("Forbidden", 403)

    # device check
    if is_mobile():
        try:
            return render_template(
                "error.html",
                message="This page is not available on mobile devices.",
                code=900
            ), 900
        except Exception as e:
            current_app.logger.warning("error.html render failed: %r", e)
            abort(403, description="This page is not available on mobile devices.")


    # normal desktop flow
    try:
        return render_template(
            "class_students.html",
            current_year=date.today().year, years = get_years(), terms = get_terms()
        )
    except Exception:
        traceback.print_exc()
        return "<pre>" + traceback.format_exc() + "</pre>", 500

    
@class_bp.route("/UploadAchievement")
@login_required
def achievement_upload():
    try:
        
        return render_template("achievement_upload.html", current_year=date.today().year, years = get_years(), terms = get_terms(), term = session.get("nearest_term"), year = session.get("nearest_year"))
    except Exception as e:
        # Print full traceback to console
        traceback.print_exc()
        # Optionally, return the traceback in the browser (only in dev!)
        return f"<pre>{traceback.format_exc()}</pre>", 500
    


# =========================
# Mutating AJAX routes
# =========================

@class_bp.route("/update_class_info", methods=["POST"])
@login_required
def update_class_info():
    if session.get("user_admin") != 1:
        return jsonify(success=False, error="Unauthorized"), 403

    data = request.get_json()
    class_id = data.get("class_id")
    class_name = data.get("class_name", "").strip()
    teacher_name = data.get("teacher_name", "").strip()

    engine = get_db_engine()
    with engine.begin() as conn:
        conn.execute(
            text("""
                EXEC FlaskHelperFunctionsSpecific 
                    @Request = :request,
                    @ClassID = :class_id,
                    @ClassName = :class_name,
                    @TeacherName = :teacher_name
            """),
            {
                "request": "UpdateClassInfo",
                "class_id": class_id,
                "class_name": class_name,
                "teacher_name": teacher_name
            }
        )
    return jsonify(success=True)


@class_bp.route('/update_competency', methods=['POST'])
@login_required
def update_competency():
    data = request.json
    nsn = data.get("nsn")
    header_name = data.get("header_name")
    status = data.get("status")
    class_id = data.get("class_id")
    term = data.get("term")
    year = data.get("year")
    debug = 0

    current_app.logger.info("📥 Incoming update_competency call")
    current_app.logger.info(f"➡️ NSN: {nsn}, Header: {header_name}, Status: {status}, Class ID: {class_id}, Term: {term}, Year: {year}")

    if None in (nsn, header_name, status, class_id, term, year):
        current_app.logger.error("❌ Missing one or more required fields")
        return jsonify({"success": False, "message": "Missing data"}), 400

    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            current_app.logger.info("🔄 Running stored procedure FlaskUpdateAchievement...")
            conn.execute(
                text("EXEC FlaskUpdateAchievement @NSN = :nsn, @Header = :header, @Value = :value, @Email = :email, @Debug = :debug"),
                {
                    "nsn": nsn,
                    "header": header_name,
                    "value": status,
                    "email": session.get("user_email"),
                    "debug": debug
                }
            )
            current_app.logger.info("✅ Stored procedure executed")

        class_cache = session.get("class_cache", {})
        updated_keys = 0
        updated_students = 0

        for key, cache in class_cache.items():
            if key.startswith(f"{class_id}_{term}_{year}_"):
                students = cache.get("student_competencies", [])
                for student in students:
                    if str(student.get("NSN")) == str(nsn):
                        current_app.logger.info(f"✏️ Updating cache for NSN {nsn}, header {header_name}")
                        student[header_name] = status
                        updated_students += 1
                updated_keys += 1

        session["class_cache"] = class_cache
        current_app.logger.info(f"✅ Cache edited for {updated_keys} key(s), {updated_students} student(s)")

        return jsonify({"success": True})
    except Exception as e:
        current_app.logger.execute("❌ Exception occurred during update_competency")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@class_bp.route("/update_scenario", methods=["POST"])
@login_required
def update_scenario():
    data = request.get_json()
    nsn = data.get("nsn")
    header = data.get("header")
    value = data.get("value")
    class_id = data.get("class_id")
    term = data.get("term")
    year = data.get("year")
    debug = 0

    current_app.logger.info(f"📥 Incoming update_scenario call")
    current_app.logger.info(f"➡️ NSN: {nsn}, Header: {header}, Value: {value}, Class ID: {class_id}, Term: {term}, Year: {year}")

    if None in (nsn, header, value, class_id, term, year):
        return jsonify(success=False, error="Missing parameters"), 400

    try:
        engine = get_db_engine()
        with engine.begin() as conn:
            current_app.logger.info("🔄 Running stored procedure FlaskUpdateAchievement...")
            conn.execute(
                text("EXEC FlaskUpdateAchievement @NSN = :nsn, @Header = :header, @Value = :value, @Email = :email, @Debug = :debug"),
                {
                    "nsn": nsn,
                    "header": header,
                    "value": value,
                    "email": session.get("user_email"),
                    "debug": debug
                }
            )
            current_app.logger.info("✅ Stored procedure executed")

        # ✏️ Inline update of session cache
        class_cache = session.get("class_cache", {})
        prefix = f"{class_id}_{term}_{year}_"
        updates = 0

        for key in list(class_cache):
            if key.startswith(prefix):
                entry = class_cache[key]
                students = entry.get("student_competencies", [])
                for student in students:
                    if isinstance(student, dict) and str(student.get("NSN")) == str(nsn):
                        current_app.logger.info(f"✏️ Updating scenario cache for NSN {nsn}, header {header} in key {key}")
                        student[header] = str(value)
                        updates += 1
                class_cache[key] = entry  # save updated version

        session["class_cache"] = class_cache
        current_app.logger.info(f"✅ Scenario cache updated in {updates} cache keys")

        return jsonify(success=True)

    except Exception as e:
        current_app.logger.execute("❌ Scenario update failed")
        traceback.print_exc()
        return jsonify(success=False, error=str(e)), 500

@class_bp.route("/class_bp/add_class", methods=["POST"])
@login_required
def add_class():
    data = request.get_json(silent=True) or {}
    moe   = data.get("moenumber")
    term  = data.get("term")
    year  = data.get("year")
    cname = (data.get("class_name") or "").strip()
    tname = (data.get("teacher_name") or "").strip()

    # Basic validation
    try:
        if not moe:
            return _json_error("Missing 'moenumber'")
        term = _require_int(term, "term")
        year = _require_int(year, "year")
        if not cname:
            return _json_error("Missing 'class_name'")
        if not tname:
            return _json_error("teacher_name is required")
    except ValueError as e:
        return _json_error(str(e))

    engine = get_db_engine()
    try:
        with engine.begin() as conn:
            # Use an OUTPUT parameter pattern to get new ClassID
            # We capture it with a temp table and select it.
            stmt = text("""
                EXEC FlaskHelperFunctionsSpecific
                    @Request = 'AddClass',
                    @MOENumber = :moe,
                    @Term = :term,
                    @CalendarYear = :year,
                    @ClassName = :cname,
                    @TeacherName = :tname
            """)
            row = conn.execute(stmt, {
                "moe": moe,
                "term": term,
                "year": year,
                "cname": cname,
                "tname": tname
            }).fetchone()

            new_id = row._mapping["NewClassID"] if row else None

        if not new_id:
            return _json_error("Class was not created (no id returned).", 500)

        return jsonify({"ok": True, "class_id": new_id, "name": cname, "teacher": tname})
    except SQLAlchemyError as e:
        return _json_error(f"Database error adding class: {str(e)}", 500)

@class_bp.route("/add_student", methods=["POST"])
@login_required
def add_student_to_class():

    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    # --- Parse body without consuming the stream ---
    # get_data() defaults cache=True, so get_json() can still read it.
    raw = request.get_data(as_text=True)

    data = request.get_json(silent=True)
    if data is None and raw:
        # Fallback: try manual JSON load (handles wrong Content-Type)
        try:
            data = json.loads(raw)
        except Exception as e:
            return _json_error("Invalid JSON", 400)
    elif data is None:
        return _json_error("Invalid JSON", 400)


    nsn = data.get("nsn")
    class_id = data.get("class_id")
    year_level = data.get("year_level")

    # normalize year_level: empty string -> None
    if year_level in ("", None):
        year_level = None

   

    if not (nsn and class_id):
        return _json_error("nsn and class_id are required")

    try:
        engine = get_db_engine()
        sql = "EXEC FlaskAddStudentToClass @NSN=:n, @ClassID=:cid, @YearLevelID=:yl"
        params = {"n": nsn, "cid": class_id, "yl": year_level}

        with engine.begin() as conn:
            conn.execute(text(sql), params)

        return jsonify({"ok": True})
    except Exception as e:
        traceback.print_exc()
        return _json_error("Failed to add student to class", 500)

@class_bp.route("/create_student_and_add", methods=["POST"])
@login_required
def create_student_and_add():
    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    d = request.get_json(silent=True) or {}
    class_id   = d.get("class_id")
    student    = d.get("student") or {}
    year_level = d.get("year_level")
    term_in    = d.get("term")   # may be None; proc can derive
    year_in    = d.get("year")   # may be None; proc can derive

    if not class_id:
        return _json_error("class_id is required")
    nsn = student.get("NSN")
    if nsn in (None, "", []):
        return _json_error("NSN is required and must be numeric")
    try:
        nsn = int(str(nsn).strip())
    except ValueError:
        return _json_error("NSN must be numeric")
    if not (student.get("FirstName") and student.get("LastName")):
        return _json_error("Student FirstName and LastName are required")
    first = (student.get("FirstName") or "").strip()
    last  = (student.get("LastName")  or "").strip()
    pref  = (student.get("PreferredName") or None)
    dob   = (student.get("DateOfBirth") or None)
    eth   = (student.get("EthnicityID") or 0)
    yl    = (year_level if year_level not in ("", None) else None)
    term  = None if term_in in ("", None) else str(term_in)
    year  = year_in  # pass through (None is fine)

    eng = get_db_engine()
    try:
        with eng.begin() as conn:            

            conn.exec_driver_sql("""
                SET NOCOUNT ON;
                DECLARE @NSN BIGINT = ?;
                EXEC  FlaskCreateStudentAddToClassAndSeed
                     @NSN=@NSN OUTPUT,
                     @FirstName=?, @LastName=?, @PreferredName=?, @DateOfBirth=?, @EthnicityID=?,
                     @ClassID=?, @CalendarYear=?, @Term=?, @YearLevelID=?,
                     @SeedScenarios=1, @SeedCompetencies=1;
            """, (nsn, first, last, pref, dob, eth, class_id, year, term, yl))

        return jsonify({"ok": True, "nsn": nsn, "class_id": class_id})

    except DBAPIError as e:
        status, friendly, sql_code = friendly_sql_error(e)
        current_app.logger.exception("create_student_and_add failed (sql=%s)", sql_code)
        return jsonify({"ok": False, "error": friendly, "sql_error": sql_code}), status
    except Exception as e:
        current_app.logger.exception("create_student_and_add failed (non-DB)")
        return jsonify({"ok": False, "error": "Unexpected error. Please try again."}), 500
    
@class_bp.route("/apply_upload", methods=["POST"])
@login_required
def apply_upload():
    """
    Body:
      { "class_id": 123, "dry_run": 1, "json_data": [ {...}, ... ] }

    Returns JSON the UI can show:
      {
        "ok": true/false,
        "status": {"ok": true/false, "message": "...", "count": N},
        "dry_run": 1,
        "term_context": {...},              # from TERM_CONTEXT
        "unexpected_students": [...],       # Info = UNEXPECTED_STUDENT
        "valid_students": [...],            # Info = VALID_STUDENT
        "competency_rows": [...],           # Info = COMPETENCY_ROWS
        "scenario_rows": [...],             # Info = SCENARIO_ROWS
        "merge_preview": []                 # kept for backward UI compatibility
      }
    """

    engine = get_db_engine()

    def _row_to_dict(cols, row):
        d = {}
        for k, v in zip(cols, row):
            # Make dates/datetimes JSON serializable
            if hasattr(v, "isoformat"):
                v = v.isoformat()
            d[k] = v
        return d

    try:
        payload  = request.get_json(silent=True) or {}
        class_id = int(payload.get("class_id") or 0)
        dry_run  = 0 
        rows     = payload.get("json_data")
        if not class_id:
            return jsonify({"ok": False, "error": "Missing class_id"}), 400
        if not isinstance(rows, list) or not rows:
            return jsonify({"ok": False, "error": "json_data must be a non-empty array"}), 400

        json_str = json.dumps(rows, ensure_ascii=False)

        # Buckets for proc outputs
        term_context         = {}
        unexpected_students  = []
        valid_students       = []
        competency_rows      = []
        scenario_rows        = []
        merge_preview        = []   # keep for UI compatibility
        status_rows_raw      = []   # if proc ever returns Ok/Message/Count again

        conn = engine.raw_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                DECLARE @j NVARCHAR(MAX) = ?,
                        @cid INT         = ?,
                        @dry BIT         = ?;
                EXEC  FlaskAchievementUpload
                     @ClassID=@cid, @JsonData=@j, @DryRun=@dry;
                """,
                (json_str, class_id, dry_run)
            )

            while True:
                if cursor.description:
                    cols = [c[0] for c in cursor.description]
                    rows_rs = cursor.fetchall()

                    # Route by "Info" label when present
                    if "Info" in cols and rows_rs:
                        info_idx = cols.index("Info")
                        info_val = str(rows_rs[0][info_idx] or "")

                        # helper: strip Info key from dicts
                        def _rows_without_info():
                            out = []
                            for r in rows_rs:
                                d = _row_to_dict(cols, r)
                                d.pop("Info", None)
                                out.append(d)
                            return out

                        if info_val == "TERM_CONTEXT":
                            # single row expected
                            d = _row_to_dict(cols, rows_rs[0])
                            d.pop("Info", None)
                            term_context = d

                        elif info_val == "UNEXPECTED_STUDENT":
                            unexpected_students.extend(_rows_without_info())

                        elif info_val == "VALID_STUDENT":
                            valid_students.extend(_rows_without_info())

                        elif info_val == "COMPETENCY_ROWS":
                            competency_rows.extend(_rows_without_info())

                        elif info_val == "SCENARIO_ROWS":
                            scenario_rows.extend(_rows_without_info())

                        else:
                            # Unknown Info label; ignore or log
                            pass

                    # Legacy status rows (Ok/Message/Count) if they ever show up
                    elif {"Ok", "Message"}.issubset(set(cols)):
                        for r in rows_rs:
                            d = _row_to_dict(cols, r)
                            status_rows_raw.append(d)

                    # Old "merge preview" (not emitted by current proc) – keep for safety
                    elif {"Action", "NSN", "CompetencyID", "YearGroupID"}.issubset(set(cols)):
                        for r in rows_rs:
                            d = _row_to_dict(cols, r)
                            d.pop("Info", None)
                            merge_preview.append(d)

                if not cursor.nextset():
                    break

            cursor.close()
            conn.commit()
        finally:
            conn.close()

        # Compute status:
        # Not OK if there are unexpected students; otherwise OK.
        if unexpected_students:
            status_obj = {
                "ok": False,
                "message": "Some uploaded students are not linked to this class.",
                "count": len(unexpected_students)
            }
            overall_ok = False
        else:
            status_obj = {
                "ok": True,
                "message": "Ready to apply.",
                "count": len(valid_students)
            }
            overall_ok = True

        # If the proc DID return Ok/Message/Count, you could override with the last row:
        if status_rows_raw:
            last = status_rows_raw[-1]
            status_obj = {
                "ok": bool(last.get("Ok")),
                "message": last.get("Message"),
                "count": int(last.get("Count") or 0),
            }
            overall_ok = status_obj["ok"]

        valid_count      = len(valid_students)
        unexpected_count = len(unexpected_students)
        total_count      = valid_count + unexpected_count

        summary = {
            "success": unexpected_count == 0,  # True when all rows are valid
            "dry_run": bool(dry_run),
            "total_rows": total_count,
            "valid_rows": valid_count,
            "unexpected_rows": unexpected_count,
        }

        return jsonify({
            "ok": overall_ok,                 # keep for compatibility; UI shouldn't throw on False
            "status": status_obj,
            "dry_run": dry_run,
            "term_context": term_context,
            "unexpected_students": unexpected_students,
            "valid_students": valid_students,
            "competency_rows": competency_rows,
            "scenario_rows": scenario_rows,
            "merge_preview": merge_preview,
            "summary": summary,              # <-- new, drives the simple banner
        })

    except pyodbc.ProgrammingError as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

# ---- API: remove student from class ----
@class_bp.route("/remove_from_class", methods=["POST"])
@login_required
def remove_from_class():
    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    d = request.get_json(force=True)
    nsn = d.get("nsn")
    class_id = d.get("class_id")
    if not (nsn and class_id):
        return _json_error("nsn and class_id are required")

    engine = get_db_engine()
    with engine.begin() as conn:
        conn.execute(
            text("EXEC FlaskRemoveStudentFromClass @NSN=:n, @ClassID=:cid, @PerformedByEmail=:em"),
            {"n": nsn, "cid": class_id, "em": session.get("user_email")}
        )
    return jsonify({"ok": True})

@class_bp.route("/move-class", methods=["POST"])
@login_required
def move_class():
  try:
    data = request.get_json(force=True) or {}
    class_id = int(data.get("class_id"))
    new_term = int(data.get("term"))
    new_year = int(data.get("year"))

    engine = get_db_engine()
    with engine.begin() as conn:
      conn.execute(
        text("""
          EXEC ChangeClassTerm
               @ClassID        = :class_id,
               @NewTerm        = :new_term,
               @NewCalendarYear = :new_year
        """),
        {
          "class_id": class_id,
          "new_term": new_term,
          "new_year": new_year,
        },
      )

    return jsonify({"ok": True})
  except Exception as e:
    current_app.logger.exception("Move class failed")
    return jsonify({"ok": False, "error": str(e)}), 500

@class_bp.route("/delete-class", methods=["POST"])
@login_required
def delete_class():
    """
    AJAX endpoint to delete a class via DeleteClass stored procedure.
    Expects JSON: {"class_id": <int>}
    Returns JSON: {"ok": true} or {"ok": false, "error": "..."}
    """

    # Parse input
    try:
        payload = request.get_json(force=True) or {}
        class_id = int(payload.get("class_id", 0))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid class ID."}), 400

    if class_id <= 0:
        return jsonify({"ok": False, "error": "Missing or invalid class ID."}), 400

    engine = get_db_engine()

    try:
        with engine.begin() as conn:
            conn.execute(
                text("EXEC dbo.DeleteClass @ClassID = :cid"),
                {"cid": class_id},
            )

        # If the proc RAISERRORs, we won't get here (exception is thrown)
        return jsonify({"ok": True})

    except Exception as e:
        # Log full stack trace server-side
        current_app.logger.exception("DeleteClass failed for ClassID=%s", class_id)

        # Try to surface a helpful message to the UI
        msg = str(e)
        # If it's a SQLAlchemy DBAPI error, the SQL Server message is often in e.orig
        try:
            if hasattr(e, "orig") and hasattr(e.orig, "args") and e.orig.args:
                msg = str(e.orig.args[0])
        except Exception:
            pass

        return jsonify({"ok": False, "error": msg}), 500
 
# =========================
# Read only API Routes
# =========================
@class_bp.route("/update_student", methods=["POST"])
@login_required
def update_student():
    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    d = request.get_json(force=True)
    # expects NSN + updated fields (FirstName, LastName, PreferredName, EthnicityID)
    engine = get_db_engine()
    with engine.begin() as conn:
        conn.execute(
            text("""
                EXEC FlaskUpdateStudent
                    @NSN=:nsn,
                    @FirstName=:fn,
                    @LastName=:ln,
                    @PreferredName=:pn,
                    @EthnicityID=:eth
            """),
            {
                "nsn": d.get("NSN"),
                "fn": d.get("FirstName"),
                "ln": d.get("LastName"),
                "pn": d.get("PreferredName"),
                "eth": d.get("EthnicityID"),
            }
        )
    return jsonify({"ok": True})


@class_bp.route("/get_schools_by_group")
@login_required
def get_schools_by_group():
    user_role = session.get("user_role")
    if user_role != "GRP":
        return jsonify([])

    group_entities = session.get("group_entities", {})
    provider_ids = [str(e["id"]) for e in group_entities.get("PRO", [])]
    funder_ids = [str(e["id"]) for e in group_entities.get("FUN", [])]
    term = request.args.get("term", type=int)
    year = request.args.get("year", type=int)

    engine = get_db_engine()
    with engine.connect() as conn:
        if provider_ids:
            csv_providers = ",".join(provider_ids)
            result = conn.execute(
                text("EXEC FlaskSchoolsByGroupProviders :ProviderList, :Term, :Year"),
                {"ProviderList": csv_providers, "Term": term, "Year": year}
            )
        elif funder_ids:
            csv_funders = ",".join(funder_ids)
            result = conn.execute(
                text("EXEC FlaskSchoolsByGroupFunders :FunderList, :Term, :Year"),
                {"FunderList": csv_funders, "Term": term, "Year": year}
            )
        else:
            return jsonify([])

        schools = [{"MOENumber": row.MOENumber, "School": row.School} for row in result]
        return jsonify(schools)

@class_bp.route("/get_schools_for_term_year")
@login_required
def get_schools_for_term_year():
    term = request.args.get("term", type=int)
    year = request.args.get("year", type=int)
    user_role = session.get("user_role")
    funder_id = session.get("user_id")
    
    if not term or not year:
        return jsonify([])

    engine = get_db_engine()
    with engine.connect() as conn:
        if user_role == "ADM":
            result = conn.execute(
                text("""
                    EXEC FlaskHelperFunctionsSpecific 
                        @Request = :Request,
                        @Term = :Term,
                        @Year = :Year
                """),
                {
                    "Request": "DistinctSchoolsByTermYear",
                    "Term": term,
                    "Year": year
                }
            )
        else:
            result = conn.execute(
                text("""
                    EXEC FlaskHelperFunctionsSpecific 
                        @Request = :Request,
                        @Term = :Term,
                        @Year = :Year,
                        @FunderID = :FunderID
                """),
                {
                    "Request": "DistinctSchoolsByTermYear",
                    "Term": term,
                    "Year": year,
                    "FunderID": funder_id
                }
            )

        rows = result.fetchall()

        return jsonify([
            {"MOENumber": row.MOENumber, "School": row.SchoolName}
            for row in rows
        ])

@class_bp.route('/get_schools_by_provider')
@login_required
def get_schools_by_provider():
    if session.get("user_role") != "PRO":
        return jsonify([])

    term = request.args.get("term", type=int)
    year = request.args.get("year", type=int)
    provider_id = session.get("user_id")

    engine = get_db_engine()
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                EXEC FlaskHelperFunctionsSpecific
                    @Request = 'SchoolsByProviderTermYear',
                    @Term = :term,
                    @Year = :year,
                    @ProviderID = :provider_id
            """), {
                "term": term,
                "year": year,
                "provider_id": provider_id
            }
        )
        schools = [dict(row._mapping) for row in result]
    return jsonify(schools)

@class_bp.route('/get_schools_by_funder')
@login_required
def get_schools_by_funder():
    user_role = session.get("user_role")
    term = request.args.get("term", type=int)
    year = request.args.get("year", type=int)

    if not term or not year:
        return jsonify([])

    engine = get_db_engine()
    with engine.connect() as conn:
        if user_role == "FUN":
            funder_id = session.get("user_id")
            result = conn.execute(
                text("""
                    EXEC FlaskHelperFunctionsSpecific
                        @Request = 'SchoolsByFunderTermYear',
                        @Term = :term,
                        @Year = :year,
                        @FunderID = :funder_id
                """), {
                    "term": term,
                    "year": year,
                    "funder_id": funder_id
                }
            )
        else:
            result = conn.execute(
                text("""
                    EXEC FlaskHelperFunctionsSpecific
                        @Request = 'SchoolsByTermYear',
                        @Term = :term,
                        @Year = :year
                """), {
                    "term": term,
                    "year": year
                }
            )
        schools = [dict(row._mapping) for row in result]
    return jsonify(schools)

@class_bp.route("/class_bp/get_classes_by_school")
@login_required
def get_classes_by_school():
    moe  = request.args.get("moe")
    term = request.args.get("term")
    year = request.args.get("year")

    try:
        if not moe:
            return _json_error("Missing 'moe'")
        term = _require_int(term, "term")
        year = _require_int(year, "year")
    except ValueError as e:
        return _json_error(str(e))

    engine = get_db_engine()
    try:
        with engine.begin() as conn:
            # Stored proc returns: ClassID, ClassName, TeacherName
            stmt = text("""
                EXEC [FlaskHelperFunctionsSpecific]
                @Request = :r,
                     @MOENumber = :moe,
                     @Term = :term,
                     @CalendarYear = :year
            """)
            rows = conn.execute(stmt, {"r":"AllClassesBySchoolTermYear","moe": moe, "term": term, "year": year}).fetchall()
        out = [
            {
                "id": r._mapping["ClassID"],
                "name": r._mapping["ClassName"],
                "teacher": r._mapping.get("TeacherName")
            }
            for r in rows
        ]
        return jsonify(out)
    except SQLAlchemyError as e:
        return _json_error(f"Database error loading classes: {str(e)}", 500)

# ---- API: classes for a school/term/year ----
@class_bp.route("/classes_for_term")
@login_required
def classes_for_term():
    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    try:
        moe = int(request.args["moe"])       # MOENumber (School ID)
        term = request.args["term"]
        year = int(request.args["year"])
    except Exception:
        return _json_error("Missing or invalid parameters")

    engine = get_db_engine()
    with engine.begin() as conn:
        # You create this proc to list classes by school/term/year
        rows = conn.execute(
            text("EXEC FlaskGetClassesForTerm @MOENumber=:m, @Term=:t, @CalendarYear=:y"),
            {"m": moe, "t": term, "y": year}
        ).fetchall()

    out = [{"id": r._mapping["ClassID"], "name": r._mapping["ClassName"]} for r in rows]
    return jsonify(out)

# ---- API: get students in a class ----
@class_bp.route("/students/<int:class_id>")
@login_required
def get_class_students(class_id):
    if not _require_moe_or_adm():
        return _json_error("Forbidden", 403)

    engine = get_db_engine()
    with engine.begin() as conn:
        rows = conn.execute(
            text("EXEC FlaskGetClassStudents @ClassID=:cid"),
            {"cid": class_id}
        ).fetchall()

    # expected columns from your proc:
    # NSN, FirstName, PreferredName, LastName, YearLevel, Ethnicity, DateOfBirth
    out = []
    for r in rows:
        m = r._mapping
        out.append({
            "NSN": m.get("NSN"),
            "FirstName": m.get("FirstName"),
            "PreferredName": m.get("PreferredName"),
            "LastName": m.get("LastName"),
            "YearLevel": m.get("YearLevelID"),
            "Ethnicity": m.get("Ethnicity"),
            "DateOfBirth": str(m.get("DateOfBirth") or "")[:10],
            "Deletable": m.get("Deletable")
        })
    return jsonify(out)

@class_bp.route("/search_students")
@login_required
def search_students():
    q        = (request.args.get("q") or "").strip()
    moe      = request.args.get("moe", type=int)
    class_id = request.args.get("class_id", type=int)

    current_app.logger.info(f"🔎 /search_students called: q='{q}', moe={moe}, class_id={class_id}")

    # Require both a school and a non-empty query
    if not (moe and q):
        current_app.logger.info("➡️  Missing moe or query → returning empty list")
        return jsonify([])

    eng = get_db_engine()
    try:
        with eng.begin() as conn:
            current_app.logger.info("➡️  Executing stored proc FlaskSearchStudentsForSchool_AllTime…")
            rows = conn.execute(
                text(
                    "EXEC FlaskSearchStudentsForSchool_AllTime "
                    "@MOENumber=:moe, @Query=:q, @ClassID=:cid"
                ),
                {"moe": moe, "q": q, "cid": class_id},
            ).fetchall()
            current_app.logger.info(f"✅ Stored proc returned {len(rows)} rows")
    except Exception as e:
        # This will surface any SQL or connection errors
        current_app.logger.exception("💥 DB call failed:", e)
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

    out = []
    for r in rows:
        m = r._mapping
        out.append(
            {
                "NSN": m.get("NSN"),
                "FirstName": m.get("FirstName"),
                "PreferredName": m.get("PreferredName"),
                "LastName": m.get("LastName"),
                "DateOfBirth": (
                    str(m.get("DateOfBirth"))[:10]
                    if m.get("DateOfBirth")
                    else ""
                ),
                "EthnicityID": m.get("EthnicityID"),
                "Ethnicity": m.get("Ethnicity"),
                "InClass": bool(m.get("InClass")),
            }
        )

    current_app.logger.info(f"➡️  Returning {len(out)} student records to client")
    return jsonify(out)


@class_bp.route("/ethnicities")
@login_required
def ethnicities():
    engine = get_db_engine()
    with engine.begin() as conn:
        rows = conn.execute(text("EXEC FlaskHelperFunctions @Request='EthnicityDropdown'")).fetchall()
    return jsonify([{"id": r._mapping["EthnicityID"], "desc": r._mapping["Description"]} for r in rows])

   
   
# =========================
# Export Routes
# =========================

@class_bp.route("/export_class_excel")
@login_required
def export_class_excel():
    try:
        engine  = get_db_engine()
        class_id = int(request.args.get("class_id"))
        term     = int(request.args.get("term"))
        year     = int(request.args.get("year"))

        _ensure_authorised_for_class(engine, class_id)
         
        meta = _get_class_meta(engine, class_id)
         

        df = _load_class_list_df(engine, class_id, term, year)
         

        if df.empty:
            df = pd.DataFrame(columns=["No results"])

        bio = excel_bytes_writer(df, sheet_name="Class List")
        fname =_safe_filename(f"{meta['SchoolName']} - {meta['ClassName']} - Class List (T{term} {year}).xlsx")
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname
        )
    except Exception:
        current_app.logger.exception("❌ export_class_excel failed:")
        # Return something visible in the browser while you’re debugging
        return jsonify({"success": False, "error": "Export failed. See server logs for details."}), 500

@class_bp.route("/export_achievements_excel", methods=["GET", "POST"])
@login_required
def export_achievements_excel():
    try:
        engine = get_db_engine()

        # Prefer JSON body; fall back to querystring/form for ids only
        payload = request.get_json(silent=True) or {}
        class_id = int(payload.get("class_id") or (request.values.get("class_id") or 0))
        term     = int(payload.get("term")     or (request.values.get("term")     or 0))
        year     = int(payload.get("year")     or (request.values.get("year")     or 0))

        _ensure_authorised_for_class(engine, class_id)
        meta = _get_class_meta(engine, class_id)

        # ---------- Build dataframe ----------
        df = None

        # POST JSON: { rows: [...] }  or { data: [...] }
        if request.method == "POST" and request.is_json:
            rows = payload.get("rows") or payload.get("data") or []
            if rows:
                # Optional: lightweight guardrails
                if not isinstance(rows, list):
                    return jsonify({"success": False, "error": "rows must be a list"}), 400
                if len(rows) > 5000:
                    return jsonify({"success": False, "error": "Too many rows"}), 413
                df = pd.DataFrame(rows)

        # Disallow giant/legacy GET with &df=... for privacy + CF limits
        if request.method == "GET" and request.args.getlist("df"):
            return jsonify({
                "success": False,
                "error": "Large GET payloads are not supported. POST a JSON body with { rows: [...] } instead."
            }), 413

        # Fallback to DB exporter (works for both GET and POST when rows weren’t provided)
        if df is None:
            df = _load_achievements_df(engine, class_id, term, year)

        if df.empty:
            df = pd.DataFrame(columns=["No results"])

        # ---------- Clean & shape ----------
        # Remove NSN
        df.drop(columns=["NSN"], errors="ignore", inplace=True)

        # Clean headers: <br> → space
        def _clean_col(c: str) -> str:
            s = str(c)
            s = re.sub(r"<br\s*/?>", " ", s, flags=re.I)
            return s.strip()

        df.rename(columns={c: _clean_col(c) for c in df.columns}, inplace=True)

        # YearLevelID → YearLevel
        if "YearLevelID" in df.columns:
            df.rename(columns={"YearLevelID": "YearLevel"}, inplace=True)

        # --- PATCH: Force column order to match UI ---
        ui_order_raw = []
        if request.method == "POST" and request.is_json:
            ui_order_raw = (payload.get("column_order") or [])

        def _norm_header_for_match(s: str) -> str:
            s = str(s)
            s = re.sub(r"<br\s*/?>", " ", s, flags=re.I)
            return s.strip()

        df_norm_map = {str(c).strip(): c for c in df.columns}

        desired_ach_cols = []
        for ui_col in ui_order_raw:
            norm = _norm_header_for_match(ui_col)
            if norm in df_norm_map:
                desired_ach_cols.append(df_norm_map[norm])

        id_cols = [c for c in ["LastName", "PreferredName", "YearLevel"] if c in df.columns]
        remaining = [c for c in df.columns if c not in id_cols + desired_ach_cols]
        df = df[id_cols + desired_ach_cols + remaining]
        # --- END PATCH ---

        # Identity first (redundant now but safe)
        id_cols = [c for c in ["LastName", "PreferredName", "YearLevel"] if c in df.columns]
        rest_cols = [c for c in df.columns if c not in id_cols]
        if id_cols:
            df = df[id_cols + rest_cols]

        # 1 → "Y", 0/NaN → "" for binary columns (non-identity)
        def _is_binary(series: pd.Series) -> bool:
            uniq = set(series.dropna().astype(str).str.strip().unique())
            return uniq.issubset({"0", "1", "0.0", "1.0"})
        for col in rest_cols:
            s = df[col]
            if _is_binary(s):
                df[col] = (
                    s.replace({1: "Y", 1.0: "Y", "1": "Y", "1.0": "Y",
                               0: "", 0.0: "", "0": "", "0.0": ""})
                     .fillna("")
                )

        # ---------- Write Excel with 2-row header ----------
        bio = io.BytesIO()
        sheet = "Achievements"

        def split_header(col_name: str) -> tuple[str, str]:
            s = str(col_name).strip()
            m = re.match(r"^(.*?)\s*(?:\((.*?)\))?\s*$", s)
            base = (m.group(1) if m else s).strip()
            in_parens = (m.group(2) if m else "").strip()
            return base, in_parens

        DATA_START_COL = 3  # D; identity are A..C

        with pd.ExcelWriter(bio, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, header=False, sheet_name=sheet, startrow=2)
            wb = writer.book
            ws = writer.sheets[sheet]

            last_col = max(0, len(df.columns) - 1)
            school = meta.get("SchoolName", "")
            klass  = meta.get("ClassName", "")
            teach  = meta.get("TeacherName", "")
            title_lines = [
                f"{school} — {klass}".strip(" —"),
                f"Teacher: {teach}" if teach else "",
                f"Term {term}, {year}",
            ]
            title_text = "\n".join([ln for ln in title_lines if ln])

            title_fmt = wb.add_format({
                "bold": True, "font_size": 12, "align": "left",
                "valign": "top", "text_wrap": True,
            })
            ws.merge_range(0, 0, 0, min(2, last_col), title_text, title_fmt)

            # Formats
            header_row1_rot = wb.add_format({
                "bold": True, "valign": "top", "align": "center",
                "text_wrap": True, "border": 1, "bg_color": "#F2F2F2",
                "rotation": 90
            })
            header_row2_h = wb.add_format({
                "bold": True, "valign": "top", "align": "center",
                "text_wrap": True, "border": 1, "bg_color": "#F2F2F2"
            })
            id_header_fmt = wb.add_format({
                "bold": True, "valign": "vcenter", "align": "left",
                "text_wrap": False, "border": 1, "bg_color": "#F2F2F2"
            })
            cell_text_fmt = wb.add_format({"valign": "bottom", "text_wrap": True})
            cell_center_fmt = wb.add_format({"valign": "vcenter", "align": "center"})

            # Header row heights
            ws.set_row(0, 120)
            ws.set_row(1, 17)
            ws.set_default_row(17)

            # A2:C2 identity headers
            for j, name in enumerate(["LastName", "PreferredName", "YearLevel"]):
                if j <= last_col:
                    ws.write(1, j, name, id_header_fmt)

            # D1.. base headers; D2.. subheaders
            for j in range(DATA_START_COL, last_col + 1):
                base, sub = split_header(df.columns[j])
                ws.write(0, j, base, header_row1_rot)
                ws.write(1, j, sub, header_row2_h)

            # Column widths
            width_map = {"LastName": 16, "PreferredName": 14, "YearLevel": 10}
            narrow_width = 6
            default_identity_width = 12
            for j, col in enumerate(df.columns):
                series = df[col]
                col_fmt = (cell_center_fmt
                           if (pd.api.types.is_numeric_dtype(series) or pd.api.types.is_bool_dtype(series))
                           else cell_text_fmt)
                width = width_map.get(col, default_identity_width if col in ["LastName","PreferredName","YearLevel"] else narrow_width)
                ws.set_column(j, j, width, col_fmt)

            # Freeze panes
            ws.freeze_panes(2, DATA_START_COL)

        bio.seek(0)
        fname = _safe_filename(
            f"{meta.get('SchoolName','')} - {meta.get('ClassName','')} - Achievements (T{term} {year}).xlsx"
        )
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname
        )

    except Exception:
        current_app.logger.exception("❌ export_achievements_excel failed:" )
        return jsonify({"success": False, "error": "Export failed. See server logs for details."}), 500

    
 # ---------- 2) POST add class (name + teacher) ----------

@class_bp.route("/preview_upload", methods=["POST"])
@login_required
def preview_upload():

    MAX_PREVIEW_ROWS = 200
    MAX_PAYLOAD_ROWS = 10000  # safety cap

    def _count_csv_rows(b, encoding="utf-8"):
        b.seek(0)
        text = io.TextIOWrapper(b, encoding=encoding, errors="ignore")
        total = sum(1 for _ in csv.reader(text))
        try:
            text.detach()
        except Exception:
            pass
        return max(total - 1, 0)

    def _count_xlsx_rows(b):
        b.seek(0)
        ws = load_workbook(b, read_only=True).active
        return ws.max_row - 1 if ws.max_row else 0

    def _count_xls_rows(b):
        b.seek(0)
        sh = xlrd.open_workbook(file_contents=b.read()).sheet_by_index(0)
        return sh.nrows - 1 if sh.nrows else 0

    FIELD_SYNONYMS = {
        "NSN": {"nsn","studentid","studentnumber","studentno","nznsn"},
        "FirstName": {"firstname","first","givenname","given"},
        "LastName": {"lastname","surname","familyname","last"},
        "PreferredName": {"preferredname","preferred","nickname","prefname"},
        "DateOfBirth": {"dateofbirth","dob","birthdate","birth","datebirth"},
        "YearLevel": {"yearlevel","year","grade","yrlevel","yeargroup"},
        "Ethnicity": {"ethnicity"},
    }
    def norm(s):
        return re.sub(r'[^a-z0-9]+', '', str(s or '').lower())

    try:
        f = request.files.get("file")
        if not f or f.filename == "":
            return jsonify({"ok": False, "error": "No file provided"}), 400

        filename = f.filename.lower()
        raw = f.read()
        buf = io.BytesIO(raw)

        # Read file
        if filename.endswith(".csv"):
            try:
                buf.seek(0); df = pd.read_csv(buf)
                total_rows = _count_csv_rows(io.BytesIO(raw))
            except UnicodeDecodeError:
                buf.seek(0); df = pd.read_csv(buf, encoding="latin-1")
                total_rows = _count_csv_rows(io.BytesIO(raw), encoding="latin-1")
        elif filename.endswith(".xlsx"):
            buf.seek(0); df = pd.read_excel(buf, engine="openpyxl")
            total_rows = _count_xlsx_rows(io.BytesIO(raw))
        elif filename.endswith(".xls"):
            buf.seek(0); df = pd.read_excel(buf, engine="xlrd")
            total_rows = _count_xls_rows(io.BytesIO(raw))
        else:
            ctype = f.mimetype or ""
            if "csv" in ctype:
                buf.seek(0); df = pd.read_csv(buf)
                total_rows = _count_csv_rows(io.BytesIO(raw))
            elif "excel" in ctype:
                buf.seek(0); df = pd.read_excel(buf, engine="openpyxl")
                total_rows = _count_xlsx_rows(io.BytesIO(raw))
            else:
                return jsonify({"ok": False, "error": "Unsupported file type"}), 400
        df = df.fillna("")

        # Detect “x-y” row that holds the group labels
        range_pat = re.compile(r'^\s*\d+\s*-\s*\d+\s*$', re.IGNORECASE)
        first_comp_col_idx = None
        header_row_idx = None
        for c_idx in range(df.shape[1]):
            col_as_str = df.iloc[:, c_idx].astype(str)
            hits = col_as_str.apply(lambda v: bool(range_pat.match(v)))
            if hits.any():
                first_comp_col_idx = c_idx
                header_row_idx = hits.idxmax()  # row index containing first match
                break

        # Build combined headers
        columns_combined = df.columns.astype(str).tolist()
        if header_row_idx is not None:
            range_row = df.iloc[header_row_idx, :].astype(str).tolist()
            combined = []
            for i, cell in enumerate(range_row):
                cell_clean = cell.strip()
                orig_col = str(df.columns[i]).strip()
                if (first_comp_col_idx is not None and i >= first_comp_col_idx and range_pat.match(cell_clean)):
                    combined.append(f"{orig_col} ({cell_clean})")
                else:
                    if cell_clean.lower().startswith("unnamed"):
                        combined.append("" if orig_col.lower().startswith("unnamed") else orig_col)
                    else:
                        combined.append(cell_clean)
            columns_combined = combined
            # data starts after that header row
            df = df.iloc[header_row_idx + 1:, :].copy()
            df.columns = columns_combined
            df = df.reset_index(drop=True)
        else:
            columns_combined = [("" if str(h).lower().startswith("unnamed") else str(h)) for h in columns_combined]
            df.columns = columns_combined

        # Non-competency columns by index threshold
        if first_comp_col_idx is None:
            non_comp_names = columns_combined[:]
            comp_start = None
        else:
            non_comp_names = columns_combined[:first_comp_col_idx]
            comp_start = int(first_comp_col_idx)

        # Map non-competency headers to canonical names
        field_mapping = {}
        for h in non_comp_names:
            n = norm(h)
            mapped = None
            if n:
                for canon, syns in FIELD_SYNONYMS.items():
                    if n == norm(canon) or any(n == s or s in n or n in s for s in syns):
                        mapped = canon
                        break
            field_mapping[h] = mapped

        # Rename to canonical
        rename_map = {h: field_mapping[h] for h in non_comp_names if field_mapping[h] and field_mapping[h] != h}
        original_headers_map = {field_mapping[h]: h for h in rename_map}
        if rename_map:
            df.rename(columns=rename_map, inplace=True)
            field_mapping = {(rename_map.get(k, k)): v for k, v in field_mapping.items()}
            non_comp_names = [rename_map.get(h, h) for h in non_comp_names]

        # Build preview
        preview_df = df.head(MAX_PREVIEW_ROWS).copy()
        columns = list(preview_df.columns)
        rows = preview_df.astype(object).values.tolist()

        # Build payload for stored proc (cap length)
        full_records = df.to_dict(orient="records")
        total_payload = len(full_records)
        if total_payload > MAX_PAYLOAD_ROWS:
            full_records = full_records[:MAX_PAYLOAD_ROWS]

        # NOTE: keep as a JSON array; front-end sends this as-is
        payload_json = full_records

        return jsonify({
            "ok": True,
            "columns": columns,
            "columns_combined": columns,
            "rows": rows,
            "total_rows": int(total_rows),
            "sample_rows": int(len(rows)),
            "competency_starts_at": comp_start,
            "non_competency_columns": non_comp_names,
            "field_mapping": field_mapping,
            "original_headers_map": original_headers_map,
            "payload_json": payload_json,
            "payload_capped": total_payload > len(full_records),
            "payload_rows": len(full_records),
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)}), 500

@class_bp.route("/SchoolClasses", methods=["GET", "POST"])
@login_required
def moe_classes():
    return redirect(url_for("class_bp.filter_classes"))


@class_bp.route("/Classes", methods=["GET", "POST"])
@login_required
def funder_classes():
    return redirect(url_for("class_bp.filter_classes"))


@class_bp.route("/ProviderClasses", methods=["GET", "POST"])
@login_required
def provider_classes():
    return redirect(url_for("class_bp.filter_classes"))